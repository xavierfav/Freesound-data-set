from openpyxl import load_workbook
import json
import numpy as np


# wb is your workbook
wb = load_workbook('easy_categories.xlsx')

# list of dictionary oer class, to store the examples and put it into json
list_easy_categories =[]

# get the content of a sheet
ws = wb.get_sheet_by_name('final_selection_v1')
for row in ws.iter_rows('C{}:C{}'.format(2, ws.max_row)):
    for cell in row:
	if cell.value != None:
            cat_id = str(row[0].value)
            list_easy_categories.append(cat_id)

set_easy_categories = set(list_easy_categories)
            
# --------------------------------------------------
with open('ontology/ontology_crowd.json') as fd:
    onto_data = json.load(fd)

for o in onto_data:
    if o['id'] in set_easy_categories:
        o['beginner_category'] = True
    else:
        o['beginner_category'] = False


#  dump updated list to json
json.dump(onto_data, open('ontology/ontology_crowd_new.json','w'), indent=4)
