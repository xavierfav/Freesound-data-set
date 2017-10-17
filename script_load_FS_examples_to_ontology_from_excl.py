from openpyxl import load_workbook
import json
import numpy as np


# wb is your workbook
wb = load_workbook('Freesound examples for FSD.xlsx')

# list of dictionary oer class, to store the examples and put it into json
list_examples =[]

# get the content of a sheet
ws = wb.get_sheet_by_name('Sheet1')
for row in ws.iter_rows('B{}:B{}'.format(4, ws.max_row)):
    for cell in row:
        # if cell contains an example
        if cell.value != None:
            cat_id = str(row[0].value)
            cat_examples = []
            
            for idx in range(10):
                if ws.cell(row=row[0].row, column=4+idx).value  != None:
                    try:
                        cat_examples.append(int(ws.cell(row=row[0].row, column=4+idx).value))
                    except:
                        pass

            # create dict with 2 keys
            cat_dict = {}
            cat_dict['id'] = cat_id
            cat_dict['positive_examples_FS'] = cat_examples

            # store the dict in a list
            list_examples.append(cat_dict)

            
# --------------------------------------------------
with open('ontology/ontology_preCrowd_temp.json') as fd:
    onto_data = json.load(fd)

for categ in list_examples:
    for ii in np.arange(len(onto_data)):
        if categ['id'] == onto_data[ii]['id']:
            onto_data[ii]['positive_examples_FS'] = categ['positive_examples_FS']
            break

#  dump updated list to json
json.dump(onto_data, open('ontology/ontology_preCrowd_temp.json','w'), indent=4)
