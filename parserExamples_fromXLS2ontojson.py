
# from openpyxl import Workbook
from openpyxl import load_workbook

import json
import numpy as np


# import openpyxl



# wb is your workbook
wb = load_workbook('IQA comments.xlsx')
# wb = openpyxl.load_workbook('C:/workbook.xlsx')


# print wb.get_sheet_names()

# get a sheet n print a cell
# sheet_ranges1 = wb["Sheet1"]
# print (sheet_ranges1.cell(row=30,column=5).internal_value)

# list of dictionary oer class, to store the examples and put it into json
listofExamples =[]
count=0


# get the content of a sheet
ws = wb.get_sheet_by_name('Sheet1')
# for row in ws.iter_rows('E{}:E{}'.format(ws.min_row,ws.max_row)):
for row in ws.iter_rows('E{}:E{}'.format(30, ws.max_row)):
    for cell in row:
        # print cell.value

        # if cell contains an example, do my thing
        if cell.value != None:
            count +=1

            # get n print category NAME********************

            # catName = ws.cell(row=row[0].row,column=1).internal_value
            catName = str(ws.cell(row=row[0].row, column=1).value)
            # print catName

            # get n print category EXAMPLES********************

            catExamples = []
            catExamples.append(int(cell.value))
            # print cell.value
            # search for more examples
            if ws.cell(row=row[0].row, column=6).value  != None:
                catExamples.append(int(ws.cell(row=row[0].row, column=6).value))
                if ws.cell(row=row[0].row, column=7).value  != None:
                    catExamples.append(int(ws.cell(row=row[0].row, column=7).value))

            # print catExamples

            # create dict with 2 keys
            catDict = {}
            catDict['name'] = catName
            catDict['positive_examples_FS'] = catExamples
            print catDict

            # store the dict in a list
            listofExamples.append(catDict)

print count
# store listofExamples in pickle?



# --------------------------------------------------

# load ontology, with field positive_examples_FS
with open('ontology_postIQA_3103.json') as fd:
    onto_data = json.load(fd)

# onto = json.load(open('ontology_postIQA_3103.json','rb'))
noMatchList =[]


for categ in listofExamples:
    print categ['name']

    # search target name in onto_data
    # ii=0
    # while categ['name'] != onto_data[ii]['name'] and ii < len(onto_data):
    #     ii += 1

    for ii in np.arange(len(onto_data)):
        if categ['name'].lower() == onto_data[ii]['name'].lower():
            print "----------match in " + str(ii)
            # edit examples
            onto_data[ii]['positive_examples_FS'] = categ['positive_examples_FS']
            print onto_data[ii]
            break

        if ii == len(onto_data)-1 and categ['name'].lower() != onto_data[ii]['name'].lower():
            print "----------couldnt find match AT ALL for " + categ['name']
            # save it in noMatchList
            noMatchList.append(categ['name'])


# here: most of them has been parsed



# left: JP
# review noMatchList and edit excel
# dum json in a nice structured way
#


#  dump updated list to json
json.dump(onto_data, open('ontology_postIQA_0304Examples_tmp.json','w'))




# JP: fix your part



# other*************************************************************

# wb1.save('B.xlsx')

# for row in wb.iter_rows(min_row=30, max_row=32, min_col=5, max_col=7 ):
#     for cell in row:
#         print(cell)

