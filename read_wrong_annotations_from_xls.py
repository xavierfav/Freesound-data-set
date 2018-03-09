

from openpyxl import load_workbook

# wb is your workbook
wb = load_workbook('kaggle3/Categories.xlsx')

fsids_to_remove_per_class = {}
count_nb_categories = 0

# get the content of a sheet
ws = wb.get_sheet_by_name('Sheet1')

# loop to search for ids to remove, for every category. position of cells is hardcoded
for row in ws.iter_rows('I{}:I{}'.format(13, ws.max_row)):
    for cell in row:

        # only if there is any fs_id to remove for each category
        if cell.value:
            count_nb_categories += 1

            # fetch cat name and id
            cat_name = str(ws.cell(row=row[0].row, column=1).value)
            cat_id = str(ws.cell(row=row[0].row, column=2).value)

            # fetch ids to remove, within a hardcoded range of columns in the current_row
            fsids_to_remove = []
            for current_row in ws.iter_rows(min_row=row[0].row, min_col=9, max_row=row[0].row, max_col=11):
                for current_cell in current_row:
                    if current_cell.value:
                        fsids_to_remove.append(int(current_cell.value))

            print(cat_name,fsids_to_remove)
            fsids_to_remove_per_class[cat_id] = fsids_to_remove


print('Removing sounds from excel in %d categories' % count_nb_categories)
print(fsids_to_remove_per_class)