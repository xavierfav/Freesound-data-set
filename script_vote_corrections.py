"""
This script reads fs_ids from Excel, which have been manually reviewed.
The result of th review is that their PP votes are not correct and hence
these votes must be deleted (for their corresponding category)

Output: dictionary 'ASO_id':[fs_id, fs_id, ...]
"""

import json
import numpy as np
import copy
import matplotlib.pyplot as plt
import os
import sys
import time
import itertools
# import xlsxwriter
import freesound
from openpyxl import load_workbook
from random import shuffle
import pickle

# ===================================================================

def get_sounds_from_pack(sound_id_target):
    """
    this function is an edition of the original one in script_kaggle3_dataset.py
    """
    client = freesound.FreesoundClient()
    client.set_token("eaa4f46407adf86c35c5d5796fd6ea8b05515dca", "token")

    # given a sound_id_target which pack we want to omit, retrieve pack_id
    sound = client.get_sound(sound_id_target)
    pack_id = int(sound.pack.split('/')[-2])
    print pack_id, sound.pack_name

    # retrieve the list of sound ids in the pack
    pack = client.get_pack(pack_id)
    pack_sounds = pack.get_sounds(fields="id,name,username", page_size=150)

    list_ids_pack_sounds = [sound.id for sound in pack_sounds]
    list_ids_pack_sounds.sort()

    # confirm
    print 'all the %d sounds in the pack: %s, of user %s' % (len(list_ids_pack_sounds), pack.name, pack.username)
    for idx, sound in enumerate(pack_sounds):
        print (idx + 1), sound.id, sound.name

    return list_ids_pack_sounds


def get_sounds_to_remove_from_excel(excel_file):
    """
    this function is an edition of the original one in script_kaggle3_dataset.py
    """
    # print('\nRemoving the following sounds from HQ, according to Excel %s, after review of HQ:' % excel_file)

    wb = load_workbook(excel_file)

    fsids_to_remove_per_class = {}
    count_nb_categories = 0

    # get the content of a sheet
    ws = wb.get_sheet_by_name('Sheet1')

    # loop to search for ids to remove, for every category. position of cells is hardcoded
    # categories start in row 3 (see format())
    # the sounds to remove start from column I (9th), and go up to T (20th)
    for row in ws.iter_rows('I{}:I{}'.format(3, ws.max_row)):
        for cell in row:

            # only if there is any fs_id to remove for each category
            if cell.value:
                count_nb_categories += 1

                # fetch cat name and id
                cat_name = str(ws.cell(row=row[0].row, column=1).value)
                cat_id = str(ws.cell(row=row[0].row, column=2).value)

                # fetch ids to remove, within a hardcoded range of columns in the current_row
                fsids_to_remove = []
                for current_row in ws.iter_rows(min_row=row[0].row, min_col=9, max_row=row[0].row, max_col=20):
                    for current_cell in current_row:
                        if current_cell.value:
                            fsids_to_remove.append(int(current_cell.value))

                print(cat_name, fsids_to_remove)
                fsids_to_remove_per_class[cat_id] = fsids_to_remove

    print('Removing sounds from excel in %d categories' % count_nb_categories)
    print(fsids_to_remove_per_class)
    return fsids_to_remove_per_class


def get_sounds_to_moveToLQ_from_excel(excel_file):
    """
    this function is an edition of the original one in script_kaggle3_dataset.py
    """
    # print('\nMoving the following sounds from HQ to LQ, according to Excel %s, after review of HQ:' % excel_file)

    wb = load_workbook(excel_file)

    fsids_to_moveToLQ_per_class = {}
    count_nb_categories = 0

    # get the content of a sheet
    ws = wb.get_sheet_by_name('dump March_09_Friday')

    # loop to search for ids to move to LQ, for every category. position of cells is hardcoded
    # categories start in row 3 (see format())
    # the sounds to move start from column L (12th), and go consecutively up to AH (34th)
    for row in ws.iter_rows('L{}:L{}'.format(3, ws.max_row)):
        for cell in row:

            # only if there is any fs_id to move for each category
            if cell.value:
                count_nb_categories += 1

                # fetch cat name and id
                cat_name = str(ws.cell(row=row[0].row, column=1).value)
                cat_id = str(ws.cell(row=row[0].row, column=2).value)

                # fetch ids to move, within a hardcoded range of columns in the current_row
                fsids_to_move = []
                for current_row in ws.iter_rows(min_row=row[0].row, min_col=12, max_row=row[0].row, max_col=34):
                    for current_cell in current_row:
                        if current_cell.value:
                            fsids_to_move.append(int(current_cell.value))

                print(cat_name, fsids_to_move)
                fsids_to_moveToLQ_per_class[cat_id] = fsids_to_move

    print('Moving sounds from HQ to LQ in %d categories' % count_nb_categories)
    print(fsids_to_moveToLQ_per_class)
    return fsids_to_moveToLQ_per_class


# ========================

FOLDER_DATA = 'kaggle3/'

try:
    # load json with ontology, to map aso_ids to understandable category names
    with open(FOLDER_DATA + 'json/ontology.json') as data_file:
        data_onto = json.load(data_file)
except:
    raise Exception('ADD AN ONTOLOGY JSON FILE TO THE FOLDER ' + FOLDER_DATA + 'json/')

# data_onto is a list of dictionaries
# to retrieve them by id: for every dict o, we create another dict where key = o['id'] and value is o
data_onto_by_id = {o['id']: o for o in data_onto}

""" # remove PP votes for fs_ids from specific categories *************************************"""

# given a sound id which pack we want to omit, retrieve all sound ids from the pack and remove them
# class: Squeak, ASO:'/m/07q6cd_'
# pack: Stomach, fsid_target = 177365, 40 samples
# pack: Human Chipmunk, fsid_target = 168135, 28 samples

fsids_to_remove_squeak = []
fsids_target = [177365, 168135]
for fsid_target in fsids_target:
    fsids_to_remove_squeak.extend(get_sounds_from_pack(fsid_target))

# get sounds to remove after final manual review of HQ
fsids_to_removePPvotes_per_class = get_sounds_to_remove_from_excel('kaggle3/Categories5.xlsx')
fsids_to_removePPvotes_per_class['/m/07q6cd_'].extend(fsids_to_remove_squeak)


# save to json
# with open('fsids_to_removePPvotes_per_class.pickle', 'wb') as handle:
#     pickle.dump(fsids_to_removePPvotes_per_class, handle)


# json.dump(fsids_to_removePPvotes_per_class, open(FOLDER_DATA + 'fsids_to_removePPvotes_per_class.json', 'w'))


""" the following sounds have PP votes but they should not due to several reasons:***********"""
# -they have several sound categories (most of em)
# -they have loud background noise (a few of em)
# -they only deserve one label, ie one category, but they have substantial FX like delay or reverb 5%
# -they only deserve one label, ie one category, but they are very strange. 5%
# the latter two do not deserve PNP (by definition), They should be in U
# but we dont know which ones are they, and they make more harm in PP, i guess

# get sounds to move from HQ to LQ after final manual review of HQ
fsids_to_changePPvotes_toPNP_per_class = get_sounds_to_moveToLQ_from_excel('kaggle3/Categories5.xlsx')


# json.dump(fsids_to_changePPvotes_toPNP_per_class, open(FOLDER_DATA + 'fsids_to_changePPvotes_toPNP_per_class.json', 'w'))

#
# # if the cat_id has children, we are dealing with a populated parent. sounds to move could be in the children too
# # retrieve the children and repeat transfer for any of them
# if data_onto_by_id[cat_id]['child_ids']:
#     print('\nMoving also sounds in the children of %s' % data_onto_by_id[cat_id]['name'])
#
#     for child_id in data_onto_by_id[cat_id]['child_ids']:
#
#


""" 
from the previous dict (fsids_to_changePPvotes_toPNP_per_class),
most of them are PNP, and from these, most of them are PNP due to several sources/categories.
Let us choose them for testing the beta version of the generation task.
If people do it good, we have annotations for FSD11k 
NOTE: 
removing all children from Music, as they usually dont have multiple sources; just some FX or weird stuff
"""

fsids_for_generation_task_fromFSD11k = copy.deepcopy(fsids_to_changePPvotes_toPNP_per_class)
count_sounds = 0
for cat_id, fs_ids in fsids_to_changePPvotes_toPNP_per_class.iteritems():

    child_id = cat_id
    # finds the top level category in the hierarchy, by analyzing consecutively the parents
    while True:
        print(data_onto_by_id[child_id]['name'])
        # note: if there are multiple parents, we return all of them, but only consider the first one
        # (Chime has this, but  by chance, the first one leads to Music)
        parent = [node_id for node_id in data_onto_by_id if child_id in data_onto_by_id[str(node_id)]['child_ids']]
        if not parent:
            # we've reached the top of the hierarchy
            break
        else:
            child_id = parent[0]
    print('Family: %s' % data_onto_by_id[child_id]['name'])

    if child_id == '/m/04rlf':
        # if the cat_id belongs to the 'Music' family
        print('-Removing %s\n' % data_onto_by_id[cat_id]['name'])
        del fsids_for_generation_task_fromFSD11k[cat_id]
    else:
        count_sounds += len(fs_ids)


print("\n\nNumber of sounds: %d" % count_sounds)
print("Number of cats: %d" % len(fsids_for_generation_task_fromFSD11k))

for idx, cat_id in enumerate(fsids_for_generation_task_fromFSD11k):
    print idx+1, data_onto_by_id[cat_id]['name']

json.dump(fsids_for_generation_task_fromFSD11k, open(FOLDER_DATA + 'fsids_for_generation_task_fromFSD11k.json', 'w'))


a = 9



