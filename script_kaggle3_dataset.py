import json
import numpy as np
import copy
import matplotlib.pyplot as plt
import os
import sys
import time
import itertools
# import xlsxwriter
import freesound
from openpyxl import load_workbook

FOLDER_DATA = 'kaggle3/'

#### DEFINE CONSTRAIN HERE ###
MINLEN = 0.3  # duration
MAXLEN = 30.0
MIN_VOTES_CAT = 70  # minimum number of votes per category to produce a QE.
# maybe useless cause all have more than 72 votes (paper)

# MIN_HQ = 40  # minimum number of sounds with HQ labels per category. this was used for dump March09
MIN_HQ = 64  # minimum number of sounds with HQ labels per category, to include Bass drum but not others

# MIN_LQ = 75  # minimum number of sounds  with LQ labels per category
MIN_HQdev_LQ = 95  # minimum number of sounds between HQ and LQ labels per category
# goal is 90 ultimately, we put a bit more cause we will discar some more samples that bear more than one label

PERCENTAGE_DEV = 0.7 # split 70 / 30 for DEV / EVAL
# PERCENTAGE_DEV = 0.625 # split 62.5 / 27.5 for DEV / EVAL

# MIN_QE = 0.68  # minimum QE to accept the LQ as decent. this was used for dump March09
MIN_QE = 0.647  # minimum QE to accept the LQ as decent, to include Bass drum

FLAG_BARPLOT = False
FLAG_BOXPLOT = False
FLAG_BARPLOT_PARENT = False
"""load initial data with votes, clip duration and ontology--------------------------------- """
'''------------------------------------------------------------------------------------------'''

# this the result of the mapping from FS sounds to ASO.
# 268k sounds with basic metadata and their corresponding ASO id.
# useful to get the duration of every sound
try:
    with open(FOLDER_DATA + 'json/FS_sounds_ASO_postIQA.json') as data_file:
        data_mapping = json.load(data_file)
except:
    raise Exception(
        'CHOOSE A MAPPING FILE AND ADD IT TO ' + FOLDER_DATA + 'json/ FOLDER (THE FILE INCLUDE DURATION INFORMATION NEEDED)')

# load json with votes, to select only PP and PNP
# try:
#     with open(FOLDER_DATA + 'json/votes_sounds_annotations.json') as data_file:
#         data_votes_raw = json.load(data_file)
# except:
#     raise Exception('ADD THE FILE CONTAINING THE VOTES (list of dict "value", "freesound_sound_id", "node_id") AND ADD IT TO THE FOLDER ' + FOLDER_DATA +'json/')

#
#
try:
    # load json with ontology, to map aso_ids to understandable category names
    with open(FOLDER_DATA + 'json/ontology.json') as data_file:
        data_onto = json.load(data_file)
except:
    raise Exception('ADD AN ONTOLOGY JSON FILE TO THE FOLDER ' + FOLDER_DATA + 'json/')

# data_onto is a list of dictionaries
# to retrieve them by id: for every dict o, we create another dict where key = o['id'] and value is o
data_onto_by_id = {o['id']: o for o in data_onto}


try:
    # load json with ontology, to map aso_ids to understandable category names
    # with open(FOLDER_DATA + 'json/votes_dumped_2018_Jan_22.json') as data_file:
    # with open(FOLDER_DATA + 'json/votes_dumped_2018_Feb_26.json') as data_file:
    # so far we were including in the data_votes_raw:
    # the trustable votes and the non trustable (verification clips not met)
    # from March1, we include only trustable
    # with open(FOLDER_DATA + 'json/votes_dumped_2018_Mar_01.json') as data_file:
    # with open(FOLDER_DATA + 'json/votes_dumped_2018_Mar_02.json') as data_file:
    with open(FOLDER_DATA + 'json/votes_dumped_2018_Mar_09.json') as data_file:
        data_votes_raw = json.load(data_file)
except:
    raise Exception('ADD AN ONTOLOGY JSON FILE TO THE FOLDER ' + FOLDER_DATA + 'json/')

# data_votes_raw is a dict where every key is a cat
# the value of every cat is a dict, that contains 5 keys: PP, PNP, NP, U, candidates
# the corresponding values are a list of Freesound ids
#
#
#
#

"""functions --------------------------------- """
'''------------------------------------------------------------------------------------------'''


def check_GT(group, fsid, catid, vote_groups, fsids_assigned_cat, data_sounds):
    # check if fsid has GT within a given group (PP,PNP,NP,U) of a category given by catid
    # if it does, add it to assigned fsids and send it to the corresponding group in data_sounds
    assigned = False
    if vote_groups[group].count(fsid) > 1:
        data_sounds[catid][group].append(fsid)
        fsids_assigned_cat.append(fsid)
        assigned = True
    return data_sounds, fsids_assigned_cat, assigned


def map_votedsound_2_disjointgroups_wo_agreement(fsid, catid, vote_groups, fsids_assigned_cat, data_sounds,
                                                 error_mapping_count_cat, count_risky_PP):
    # map the voted sound to a disjoint group  without inter-annotator agreement
    # using set of arbitrary rules that cover all possible options
    # being demanding now. only sending to PP when we are sure

    # retrieve votes for fsid.
    # we know there is only one in PP
    votes = []
    if fsid in vote_groups['PP']:
        votes.append(1.0)
    if fsid in vote_groups['PNP']:
        votes.append(0.5)
    if fsid in vote_groups['U']:
        votes.append(0.0)
    if fsid in vote_groups['NP']:
        votes.append(-1.0)

    # votes has all the votes for fsid. let us take decisions

    # trivial cases where there is only one single vote by one annotator
    if 1.0 in votes and 0.5 not in votes and -1.0 not in votes and 0.0 not in votes:
        # the only case where a sound is sent to PP without inter-annotator agreement
        data_sounds[catid]['PP'].append(fsid)
        fsids_assigned_cat.append(fsid)
    elif 1.0 not in votes and 0.5 in votes and -1.0 not in votes and 0.0 not in votes:
        # data_sounds[catid]['PNP'].append(fsid)
        # single vote of PNP may be a bit unreliable. safer to send it to U group
        # thus it goes to LQ (and not LQprior)
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)
    elif 1.0 not in votes and 0.5 not in votes and -1.0 in votes and 0.0 not in votes:
        data_sounds[catid]['NP'].append(fsid)
        fsids_assigned_cat.append(fsid)
    elif 1.0 not in votes and 0.5 not in votes and -1.0 not in votes and 0.0 in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)


    # rest of 11 cases. placing first the PP options in case there is some error.

    # 8 PP and PNP
    elif 1.0 in votes and 0.5 in votes and -1.0 not in votes and 0.0 not in votes:
        data_sounds[catid]['PNP'].append(fsid)
        fsids_assigned_cat.append(fsid)
        count_risky_PP += 1
    # 9 PP and PNP and U
    elif 1.0 in votes and 0.5 in votes and -1.0 not in votes and 0.0 in votes:
        data_sounds[catid]['PNP'].append(fsid)
        fsids_assigned_cat.append(fsid)
        count_risky_PP += 1


    # 1: NP and U
    elif 1.0 not in votes and 0.5 not in votes and -1.0 in votes and 0.0 in votes:
        data_sounds[catid]['NP'].append(fsid)
        fsids_assigned_cat.append(fsid)
    # 2: PNP and U
    elif 1.0 not in votes and 0.5 in votes and -1.0 not in votes and 0.0 in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)
    # 3: PNP and NP
    elif 1.0 not in votes and 0.5 in votes and -1.0 in votes and 0.0 not in votes:
        data_sounds[catid]['NP'].append(fsid)
        fsids_assigned_cat.append(fsid)
    # 4: PNP and NP and U
    elif 1.0 not in votes and 0.5 in votes and -1.0 in votes and 0.0 in votes:
        data_sounds[catid]['NP'].append(fsid)
        fsids_assigned_cat.append(fsid)


    # 5: PP and U
    elif 1.0 in votes and 0.5 not in votes and -1.0 not in votes and 0.0 in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)
    # 6: PP and NP
    elif 1.0 in votes and 0.5 not in votes and -1.0 in votes and 0.0 not in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)
    # 7: PP and NP and U
    elif 1.0 in votes and 0.5 not in votes and -1.0 in votes and 0.0 in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)



    # 10: PP and PNP and NP
    elif 1.0 in votes and 0.5 in votes and -1.0 in votes and 0.0 not in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)
    # 11: PP and PNP and NP and U
    elif 1.0 in votes and 0.5 in votes and -1.0 in votes and 0.0 in votes:
        data_sounds[catid]['U'].append(fsid)
        fsids_assigned_cat.append(fsid)

    else:
        # print('\n something unexpetected happened in the mapping********************* \n')
        error_mapping_count_cat += 1
        # sys.exit('something unexpetected happened in the mapping!')

    return data_sounds, fsids_assigned_cat, error_mapping_count_cat, count_risky_PP


# turn interactive mode on
plt.ion()


# SPLITFIGS = False
# SLAVE_BOXPLOT_DURATIONS = True



def plot_boxplot(data, x_labels, fig_title, y_label):
    fig = plt.figure()
    ax = fig.add_subplot(111)
    # for tick in ax.get_xticklabels():
    #     tick.set_rotation(45)
    plt.xticks(fontsize=8, rotation=45)
    bp = ax.boxplot(data, patch_artist=True)
    ## change color and linewidth of the whiskers
    for whisker in bp['whiskers']:
        whisker.set(color='#7570b3', linewidth=1)

    for flier in bp['fliers']:
        flier.set(marker='.', color='#e7298a', alpha=0.5)

    ## change color and linewidth of the medians
    for median in bp['medians']:
        median.set(color='#ef0707', linewidth=3)

    ax.set_xticklabels(x_labels)
    plt.ylabel(y_label)
    plt.title(fig_title)
    plt.grid(True)
    plt.show()
    # plt.pause(0.001)


def plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda):
    ind = np.arange(len(data_bottom))  # the x locations for the groups
    width = 0.5  # the width of the bars: can also be len(x) sequence
    # axes = [-0.5, len(data_bottom), 0, 170]

    fig = plt.figure()
    ax = fig.add_subplot(111)
    p1 = plt.bar(ind, data_bottom, width, color='b')
    p2 = plt.bar(ind, data_up, width, bottom=data_bottom, color='r')
    plt.xticks(fontsize=8, rotation=45)
    plt.xticks(ind, x_labels)
    plt.ylabel(y_label)
    plt.title(fig_title)
    # plt.yticks(np.arange(0, 81, 10))
    plt.legend((p1[0], p2[0]), legenda)
    # plt.axis(axes)
    ax.yaxis.grid(True)
    # plt.grid(True)
    plt.show()


def compute_median(data):
    nb_samples = []
    for id, group in data.iteritems():
        nb_samples.append(np.ceil(PERCENTAGE_DEV * len(group['HQ'])) + len(group['LQ']))
    # print 'Estimated Median of number of DEV samples per category: ' + str(np.median(nb_samples))
    # print()
    return nb_samples


def create_var_barplot(data_set, data_onto_by_id):
    # create variable with data for barplotting - function
    var_plot = []
    for catid, groups in data_set.iteritems():
        cat_plot = {}
        cat_plot['nbHQ_dev'] = np.ceil(PERCENTAGE_DEV * len(groups['HQ']))
        cat_plot['nbLQ'] = len(groups['LQ'])
        cat_plot['nbtotal_tr'] = cat_plot['nbHQ_dev'] + cat_plot['nbLQ']
        cat_plot['name'] = data_onto_by_id[str(catid)]['name']
        var_plot.append(cat_plot)
    return var_plot


def get_sounds_from_pack(sound_id_target):

    client = freesound.FreesoundClient()
    client.set_token("eaa4f46407adf86c35c5d5796fd6ea8b05515dca", "token")

    # given a sound_id_target which pack we want to omit, retrieve pack_id
    sound = client.get_sound(sound_id_target)
    pack_id = int(sound.pack.split('/')[-2])
    print pack_id, sound.pack_name

    # retrieve the list of sound ids in the pack
    pack = client.get_pack(pack_id)
    pack_sounds = pack.get_sounds(fields="id,name,username", page_size=150)

    list_ids_pack_sounds = [sound.id for sound in pack_sounds]
    list_ids_pack_sounds.sort()

    # confirm
    print 'all the %d sounds in the pack: %s, of user %s' % (len(list_ids_pack_sounds), pack.name, pack.username)
    for idx, sound in enumerate(pack_sounds):
        print (idx + 1), sound.id, sound.name

    return list_ids_pack_sounds


def get_sounds_to_remove_from_excel(excel_file):

    print('\nRemoving the following sounds from HQ, according to Excel %s, after review of HQ:' % excel_file)

    wb = load_workbook(excel_file)

    fsids_to_remove_per_class = {}
    count_nb_categories = 0

    # get the content of a sheet
    ws = wb.get_sheet_by_name('Sheet1')

    # loop to search for ids to remove, for every category. position of cells is hardcoded
    # categories start in row 3 (see format())
    # the sounds to remove start from column I (9th), and go up to T (20th)
    for row in ws.iter_rows('I{}:I{}'.format(3, ws.max_row)):
        for cell in row:

            # only if there is any fs_id to remove for each category
            if cell.value:
                count_nb_categories += 1

                # fetch cat name and id
                cat_name = str(ws.cell(row=row[0].row, column=1).value)
                cat_id = str(ws.cell(row=row[0].row, column=2).value)

                # fetch ids to remove, within a hardcoded range of columns in the current_row
                fsids_to_remove = []
                for current_row in ws.iter_rows(min_row=row[0].row, min_col=9, max_row=row[0].row, max_col=20):
                    for current_cell in current_row:
                        if current_cell.value:
                            fsids_to_remove.append(int(current_cell.value))

                print(cat_name, fsids_to_remove)
                fsids_to_remove_per_class[cat_id] = fsids_to_remove

    print('Removing sounds from excel in %d categories' % count_nb_categories)
    print(fsids_to_remove_per_class)
    return fsids_to_remove_per_class


def get_sounds_to_moveToLQ_from_excel(excel_file):

    print('\nMoving the following sounds from HQ to LQ, according to Excel %s, after review of HQ:' % excel_file)

    wb = load_workbook(excel_file)

    fsids_to_moveToLQ_per_class = {}
    count_nb_categories = 0

    # get the content of a sheet
    ws = wb.get_sheet_by_name('dump March_09_Friday')

    # loop to search for ids to move to LQ, for every category. position of cells is hardcoded
    # categories start in row 3 (see format())
    # the sounds to remove start from column L (12th), and go consecutively up to AB (28th)
    for row in ws.iter_rows('L{}:L{}'.format(3, ws.max_row)):
        for cell in row:

            # only if there is any fs_id to move for each category
            if cell.value:
                count_nb_categories += 1

                # fetch cat name and id
                cat_name = str(ws.cell(row=row[0].row, column=1).value)
                cat_id = str(ws.cell(row=row[0].row, column=2).value)

                # fetch ids to move, within a hardcoded range of columns in the current_row
                fsids_to_move = []
                for current_row in ws.iter_rows(min_row=row[0].row, min_col=12, max_row=row[0].row, max_col=28):
                    for current_cell in current_row:
                        if current_cell.value:
                            fsids_to_move.append(int(current_cell.value))

                print(cat_name, fsids_to_move)
                fsids_to_moveToLQ_per_class[cat_id] = fsids_to_move

    print('Removing sounds from excel in %d categories' % count_nb_categories)
    print(fsids_to_moveToLQ_per_class)
    return fsids_to_moveToLQ_per_class


# -------------------------end of functions-----------------------------------------------------------
# ----------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------


""" # remove unwanted sounds from specific categories before doing anything else*************************************"""

# given a sound id which pack we want to omit, retrieve all sound ids from the pack and remove them
# class: Squeak, ASO:'/m/07q6cd_'
# pack: Stomach, fsid_target = 177365, 40 samples
# pack: Human Chipmunk, fsid_target = 168135, 28 samples

fsids_to_remove_squeak = []
fsids_target = [177365, 168135]
for fsid_target in fsids_target:
    fsids_to_remove_squeak.extend(get_sounds_from_pack(fsid_target))

# # remove the sounds from data_votes_raw
data_votes = copy.deepcopy(data_votes_raw)
#
# # empty Squeak category (empty dict as value)
# data_votes['/m/07q6cd_'].clear()
#
# for key, vote_group in data_votes_raw['/m/07q6cd_'].iteritems():
#     data_votes['/m/07q6cd_'][key] = [fsid for fsid in vote_group if fsid not in fsids_to_remove_squeak]
#


# get sounds to remove after final manual review of HQ
fsids_to_remove_per_class_excel = get_sounds_to_remove_from_excel('kaggle3/Categories3.xlsx')

for cat_id, fsids_to_remove_oneclass in fsids_to_remove_per_class_excel.iteritems():
    # empty category (empty dict as value)
    data_votes[cat_id].clear()

    if cat_id == '/m/07q6cd_':
        print('Squeak case is different. join both lists to remove')
        fsids_to_remove_oneclass.extend(fsids_to_remove_squeak)

    # if cat_id == '/m/07cx4':
    #     e = 4

    count_removed_sounds_in_candidates = 0
    for key, vote_group in data_votes_raw[cat_id].iteritems():
        # removing sounds from the cat_id that appears in the excel
        data_votes[cat_id][key] = [fsid for fsid in vote_group if fsid not in fsids_to_remove_oneclass]

    count_removed_sounds_in_candidates += len(data_votes_raw[cat_id]['candidates']) - len(data_votes[cat_id]['candidates'])
    print('Removed %d sounds from %s, with id %s, out of %d sounds to remove'
          % (count_removed_sounds_in_candidates, data_onto_by_id[cat_id]['name'], cat_id, len(fsids_to_remove_oneclass)))

    # if the cat_id has children, we are dealing with a populated parent. sounds to delete could be in the children too
    # retrieve the children and repeat removal for any of them
    if data_onto_by_id[cat_id]['child_ids']:
        print('\nRemoving also sounds in the children of %s' % data_onto_by_id[cat_id]['name'])

        for child_id in data_onto_by_id[cat_id]['child_ids']:
            # empty category (empty dict as value)
            data_votes[child_id].clear()

            for key_child, vote_group_child in data_votes_raw[child_id].iteritems():
                # removing sounds from the child_id of the parent that appears in the excel
                data_votes[child_id][key_child] = [fsid for fsid in vote_group_child if fsid not in fsids_to_remove_oneclass]

            sounds_removed_child = len(data_votes_raw[child_id]['candidates']) - len(data_votes[child_id]['candidates'])
            print('---Removed %d sounds in %s, out of %d sounds to remove'
                  % (sounds_removed_child, data_onto_by_id[child_id]['name'], len(fsids_to_remove_oneclass)))

            count_removed_sounds_in_candidates += sounds_removed_child
    else:
        # sanity check ONLY for the leaves. what remains + what was to be removed = initial total
        if len(data_votes[cat_id]['candidates']) + len(fsids_to_remove_oneclass) != len(
                data_votes_raw[cat_id]['candidates']):
            print('----------------Something wrong when removing sounds from the leave %s. %d sound(s) were not removed'
                  % (data_onto_by_id[cat_id]['name'],
                     (len(data_votes[cat_id]['candidates']) + len(fsids_to_remove_oneclass) -
                      len(data_votes_raw[cat_id]['candidates']))))

        # sanity check ONLY for the leaves. how many of the files to remove were found (and hence removed)?
        count_removed_sounds_in_PP = 0
        for fs_id in fsids_to_remove_oneclass:
            if fs_id in data_votes_raw[cat_id]['PP']:
                count_removed_sounds_in_PP += 1
        # print('Removed %d sounds from PP in %s, out of %d sounds to remove'
        #       % (count_removed_sounds_in_PP, data_onto_by_id[cat_id]['name'], len(fsids_to_remove_oneclass)))
        if count_removed_sounds_in_PP != len(fsids_to_remove_oneclass):
            print('------------------We are trying to remove sounds that do not exist in category %s with id %s'
                  % (data_onto_by_id[cat_id]['name'], cat_id))

    # sanity check for ALL classes
    if count_removed_sounds_in_candidates < len(fsids_to_remove_oneclass):
        print('----------------------Something wrong when removing sounds from %s.' % (data_onto_by_id[cat_id]['name']))



""" # from data_votes to data_sounds ******************************************************************************"""
# Assign sounds to disjoint GROUPS (PP, PNP, NP, U) based on the combination of votes that they have
# Compute also QE for every category

# create copy of data_votes
data_sounds = copy.deepcopy(data_votes)
for catid, vote_groups in data_sounds.iteritems():
    data_sounds[catid]['PP'] = []
    data_sounds[catid]['PNP'] = []
    data_sounds[catid]['NP'] = []
    data_sounds[catid]['U'] = []
    data_sounds[catid]['QE'] = 0    # initialzed to 0. only if more than MIN_VOTES_CAT, we compute it

# count cases where the mapping from votes to sounds fails
error_mapping_count_cats = []

# to keep track of combinations
# PP + PNP and PP + PNP + U
count_risky_PP = 0

for catid, vote_groups in data_votes.iteritems():
    # list to keep track of assigned fsids within a category, to achieve disjoint subsets of audio samples
    fsids_assigned_cat = []
    error_mapping_count_cat = 0

    # check GT in PP
    # check GT in the rest of the groups
    # if GT does not exist, take mapping decision without inter-annotator agreement
    for fsid in vote_groups['PP']:
        # print fsid
        # search for GT in this group
        if vote_groups['PP'].count(fsid) > 1:
            if fsid not in fsids_assigned_cat:
                data_sounds[catid]['PP'].append(fsid)
                fsids_assigned_cat.append(fsid)
        else:
            # search for GT in other groups of votes
            data_sounds, fsids_assigned_cat, assigned = check_GT('PNP', fsid, catid, vote_groups, fsids_assigned_cat,
                                                                 data_sounds)
            if not assigned:
                data_sounds, fsids_assigned_cat, assigned = check_GT('U', fsid, catid, vote_groups, fsids_assigned_cat,
                                                                     data_sounds)
            if not assigned:
                data_sounds, fsids_assigned_cat, assigned = check_GT('NP', fsid, catid, vote_groups, fsids_assigned_cat,
                                                                     data_sounds)

        # no GT was found for the annotation (2 votes in the same group).
        # we must take decisions without inter-annotator agreement

        if fsid not in fsids_assigned_cat:
            # map the voted sound to a disjoint group  without inter-annotator agreement
            data_sounds, fsids_assigned_cat, error_mapping_count_cat, count_risky_PP = map_votedsound_2_disjointgroups_wo_agreement(
                fsid, catid, vote_groups, fsids_assigned_cat, data_sounds, error_mapping_count_cat, count_risky_PP)

    # check GT in PNP
    # check GT in the remaining groups
    # if GT does not exist, take mapping decision without inter-annotator agreement
    for fsid in vote_groups['PNP']:
        # print fsid

        # only if the fsid has not been assigned in previous passes
        if fsid not in fsids_assigned_cat:
            # search for GT in this group
            if vote_groups['PNP'].count(fsid) > 1:
                if fsid not in fsids_assigned_cat:
                    data_sounds[catid]['PNP'].append(fsid)
                    fsids_assigned_cat.append(fsid)
            else:
                # search for GT in the remaining groups of votes
                data_sounds, fsids_assigned_cat, assigned = check_GT('U', fsid, catid, vote_groups, fsids_assigned_cat,
                                                                     data_sounds)
                if not assigned:
                    data_sounds, fsids_assigned_cat, assigned = check_GT('NP', fsid, catid, vote_groups,
                                                                         fsids_assigned_cat, data_sounds)

            # no GT was found for the annotation (2 votes in the same group).
            # we must take decisions without inter-annotator agreement

            if fsid not in fsids_assigned_cat:
                # map the voted sound to a disjoint group  without inter-annotator agreement
                data_sounds, fsids_assigned_cat, error_mapping_count_cat, count_risky_PP = map_votedsound_2_disjointgroups_wo_agreement(
                    fsid, catid, vote_groups, fsids_assigned_cat, data_sounds, error_mapping_count_cat, count_risky_PP)

    # check GT in U
    # check GT in the remaining groups
    # if GT does not exist, take mapping decision without inter-annotator agreement
    for fsid in vote_groups['U']:
        # print fsid

        # only if the fsid has not been assigned in previous passes
        if fsid not in fsids_assigned_cat:
            # search for GT in this group
            if vote_groups['U'].count(fsid) > 1:
                if fsid not in fsids_assigned_cat:
                    data_sounds[catid]['U'].append(fsid)
                    fsids_assigned_cat.append(fsid)
            else:
                # search for GT in the remaining groups of votes
                data_sounds, fsids_assigned_cat, assigned = check_GT('NP', fsid, catid, vote_groups, fsids_assigned_cat,
                                                                     data_sounds)

            # no GT was found for the annotation (2 votes in the same group).
            # we must take decisions without inter-annotator agreement

            if fsid not in fsids_assigned_cat:
                # map the voted sound to a disjoint group  without inter-annotator agreement
                data_sounds, fsids_assigned_cat, error_mapping_count_cat, count_risky_PP = map_votedsound_2_disjointgroups_wo_agreement(
                    fsid, catid, vote_groups, fsids_assigned_cat, data_sounds, error_mapping_count_cat, count_risky_PP)

    # check GT in NP
    # check GT in the remaining groups? no need to. already done in previous passes
    # if GT does not exist, take mapping decision without inter-annotator agreement
    for fsid in vote_groups['NP']:
        # print fsid

        # only if the fsid has not been assigned in previous passes
        if fsid not in fsids_assigned_cat:
            # search for GT in this group
            if vote_groups['NP'].count(fsid) > 1:
                if fsid not in fsids_assigned_cat:
                    data_sounds[catid]['NP'].append(fsid)
                    fsids_assigned_cat.append(fsid)
            # else: no need to. already done in previous passes

            # no GT was found for the annotation (2 votes in the same group).
            # we must take decisions without inter-annotator agreement
            else:
                # if fsid not in fsids_assigned_cat:
                # map the voted sound to a disjoint group  without inter-annotator agreement
                data_sounds, fsids_assigned_cat, error_mapping_count_cat, count_risky_PP = map_votedsound_2_disjointgroups_wo_agreement(
                    fsid, catid, vote_groups, fsids_assigned_cat, data_sounds, error_mapping_count_cat, count_risky_PP)

    # store mapping error for every category
    error_mapping_count_cats.append(error_mapping_count_cat)

    # for every category compute QE here number of votes len(PP) + len(PNP) / all
    # QE should only be computed if there are more than MIN_VOTES_CAT votes. else not reliable
    if (len(vote_groups['PP']) + len(vote_groups['PNP']) + len(vote_groups['NP']) + len(
            vote_groups['U'])) >= MIN_VOTES_CAT:
        data_sounds[catid]['QE'] = (len(vote_groups['PP']) + len(vote_groups['PNP'])) / float(
            len(vote_groups['PP']) + len(vote_groups['PNP']) + len(vote_groups['NP']) + len(vote_groups['U']))

# =======================SANITY CHECKS===================================

    # sanity check: there should be no duplicated fsids within a group of data_sounds
    if (len(data_sounds[catid]['PP']) != len(set(data_sounds[catid]['PP'])) or
                len(data_sounds[catid]['PNP']) != len(set(data_sounds[catid]['PNP'])) or
                len(data_sounds[catid]['NP']) != len(set(data_sounds[catid]['NP'])) or
                len(data_sounds[catid]['U']) != len(set(data_sounds[catid]['U']))):
        # print('\n something unexpetected happened in the mapping********************* \n')
        print(catid)
        sys.exit('duplicates in data_sounds')

    # sanity check: groups in data_sounds should be disjoint
    if (list(set(data_sounds[catid]['PP']) & set(data_sounds[catid]['PNP'])) or
            list(set(data_sounds[catid]['PP']) & set(data_sounds[catid]['NP'])) or
            list(set(data_sounds[catid]['PP']) & set(data_sounds[catid]['U'])) or
            list(set(data_sounds[catid]['PNP']) & set(data_sounds[catid]['NP'])) or
            list(set(data_sounds[catid]['PNP']) & set(data_sounds[catid]['U'])) or
            list(set(data_sounds[catid]['NP']) & set(data_sounds[catid]['U']))):
        # print('\n something unexpetected happened in the mapping********************* \n')
        print(catid)
        sys.exit('data_sounds has not disjoint groups')

    # sanity check: number of sounds is equal in data_sounds (adding) and data_votes (concatenating groups and set)
    nb_sounds_data_sounds = (len(data_sounds[catid]['PP']) + len(data_sounds[catid]['PNP']) + \
                             len(data_sounds[catid]['NP']) + len(data_sounds[catid]['U']))
    all_votes = data_votes[catid]['PP'] + data_votes[catid]['PNP'] + data_votes[catid]['NP'] + data_votes[catid]['U']
    nb_sounds_data_votes = len(set(all_votes))
    if nb_sounds_data_sounds != nb_sounds_data_votes:
        # print('\n something unexpetected happened in the mapping********************* \n')
        print(catid)
        sys.exit('number of sounds is not equal in data_sounds and data_votes')

if sum(error_mapping_count_cats) > 0:
    print 'there are errors in the following number of fsids: ' + str(sum(error_mapping_count_cats))
    print(error_mapping_count_cats)

print 'Number of sounds with voting combination (PP + PNP) or (PP + PNP + U) : ' + str(count_risky_PP)
# but this number is considering all the ontology, all durations, no filter has been applied yet.

# here we have data_sounds ready



""" # from data_sounds to data_qual_sets****************************************************************************"""
# For every category, merge the disjoint groups of sounds (PP, PNP, NP, U) in two main subsets HQ and LQ.
# HQ contains only the PP group and will be denoted as 'manually verified'
# LQ contains PNP + U + unvoted sounds, and will be denoted as 'non verified'
# Additionally, we create LQprior: sounds within LQ that come from the PNP group. two reasons:
# -these sounds have been voted and hence do not depend on QE. They can always be used.
# -when filling a category with LQ, these should be prioritized.
# creating HQ and LQ with 2 versions. CASE A is preferred.

# create data_qual_sets with keys with catids and empty dicts as values
data_qual_sets = copy.deepcopy(data_votes)
for catid, groups in data_qual_sets.iteritems():
    data_qual_sets[catid].clear()

for ii in range(1):
    if ii == 0:
        # CASE A
        # HQ = PP
        # LQ = candidates - PP - NP. IT therefore includes: PNP + U + unvoted sounds
        # we use PNP as LQ, so HQ is even higher quality and avoid potential problems, eg incomplete labelling in EVAL
        # but this means we have less samples in HQ, which can reduce the number of categories
        print 'SCENARIO A: where HQ = PP and LQ = PNP + U + unvoted sounds'
        print()
        for catid, sound_groups in data_sounds.iteritems():
            data_qual_sets[catid]['HQ'] = sound_groups['PP']
            list_woPP = [item for item in sound_groups['candidates'] if item not in sound_groups['PP']]
            data_qual_sets[catid]['LQ'] = [item for item in list_woPP if item not in sound_groups['NP']]
            data_qual_sets[catid]['QE'] = sound_groups['QE']
            # sounds in LQprior are already contained in LQ
            data_qual_sets[catid]['LQprior'] = sound_groups['PNP']
    else:
        # CASE B
        # HQ = PP + PNP
        # LQ = candidates - PP - PNP - NP. IT therefore includes: U + unvoted sounds
        # we use PNP as HQ, so HQ has more heterogeneous content.
        # If there is a file with incomplete labelling in EVAL, it is not good
        # We can always leave them only for DEV
        # This means we have more samples in HQ, which can increment the number of categories
        # Maybe this is less clean but if needed...
        print 'SCENARIO B: where HQ = PP + PNP and LQ = U + unvoted sounds'
        # print()
        # for catid, sound_groups in data_sounds.iteritems():
        #     data_qual_sets[catid]['HQ'] = sound_groups['PP'] + sound_groups['PNP']
        #     list_woPP = [item for item in sound_groups['candidates'] if item not in sound_groups['PP']]
        #     list_woPP_PNP = [item for item in list_woPP if item not in sound_groups['PNP']]
        #     data_qual_sets[catid]['LQ'] = [item for item in list_woPP_PNP if item not in sound_groups['NP']]
        #     data_qual_sets[catid]['QE'] = sound_groups['QE']

    # sanity check: groups in data_qual_sets should be disjoint
    for catid in data_qual_sets:
        if list(set(data_qual_sets[catid]['HQ']) & set(data_qual_sets[catid]['LQ'])):
            # print('\n something unexpected happened in the mapping********************* \n')
            print(catid)
            sys.exit('data_qual_sets has not disjoint groups')

    # nb_samples_cats_dev = compute_median(data_qual_sets)
    # plots before strong filters: all possible categories
    # if FLAG_PLOT:
    #     # boxplot number of examples per category in dev
    #     x_labels = 'estimated dev'
    #     fig_title = 'estimated number of clips in categories of DEV set'
    #     y_label = '# of audio clips'
    #     plot_boxplot(nb_samples_cats_dev, x_labels, fig_title, y_label)
    #
    #     # create variable with data for barplotting - function
    #     var_barplot = create_var_barplot(data_qual_sets, data_onto_by_id)
    #
    #     # barplot with number of sounds per category
    #     # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
    #     idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
    #     # idx = np.argsort([data_onto_rich[tt]['nbHQ_dev'] for tt in range(len(data_onto_rich))])
    #
    #     data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
    #     data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
    #     x_labels = list(var_barplot[val]['name'] for val in idx)
    #     y_label = '# of audio clips'
    #     fig_title = 'number of DEV clips per category, split in HQdev and LQ, sorted by total'
    #     legenda = ('LQ', 'HQdev')
    #     plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)
    #


    # at this point we have: data_qual_sets, with dicts containing
    # HQ, LQ, QE for all categories (632)

    """but by reviewing the HQ part we detect some sounds that must be in LQ instead"""

    # get sounds to move from HQ to LQ after final manual review of HQ
    fsids_to_moveToLQ_per_class_excel = get_sounds_to_moveToLQ_from_excel('kaggle3/Categories3.xlsx')
    # ok till here

    for cat_id, fsids_to_move_oneclass in fsids_to_moveToLQ_per_class_excel.iteritems():

        count_moved_toLQ = 0
        print('Moving in %s' % data_onto_by_id[cat_id]['name'])
        for fs_id in fsids_to_move_oneclass:
            if fs_id in data_qual_sets[cat_id]['HQ']:
                # remove from HQ
                data_qual_sets[cat_id]['HQ'].remove(fs_id)
                # move to LQ
                data_qual_sets[cat_id]['LQ'].append(fs_id)
                # count
                count_moved_toLQ += 1

            elif not data_onto_by_id[cat_id]['child_ids']:
                # if some sound was not found, the only explanation is that it is in the children
                print('----------------Something wrong when moving sounds from HQ to LQ in the leave %s. '
                      % (data_onto_by_id[cat_id]['name']))

        print('Moved %d sounds from HQ to LQ in %s, out of %d sounds to move'
              % (count_moved_toLQ, data_onto_by_id[cat_id]['name'], len(fsids_to_move_oneclass)))

        # if the cat_id has children, we are dealing with a populated parent. sounds to move could be in the children too
        # retrieve the children and repeat transfer for any of them
        if data_onto_by_id[cat_id]['child_ids']:
            print('\nMoving also sounds in the children of %s' % data_onto_by_id[cat_id]['name'])

            for child_id in data_onto_by_id[cat_id]['child_ids']:
                count_moved_toLQ_child = 0
                for fs_id in fsids_to_move_oneclass:
                    if fs_id in data_qual_sets[child_id]['HQ']:
                        # remove from HQ
                        data_qual_sets[child_id]['HQ'].remove(fs_id)
                        # move to LQ
                        data_qual_sets[child_id]['LQ'].append(fs_id)
                        # count
                        count_moved_toLQ_child += 1

                print('---Moved %d sounds from HQ to LQ in %s, out of %d sounds to move'
                      % (count_moved_toLQ_child, data_onto_by_id[child_id]['name'], len(fsids_to_move_oneclass)))

                count_moved_toLQ += count_moved_toLQ_child

        else:
            # sanity check ONLY for the leaves. what we moved = the reading from excel
            if len(fsids_to_move_oneclass) != count_moved_toLQ:
                print('----------------Something wrong when moving sounds from HQ to LQ in the leave %s. '
                      '%d sound(s) were not moved' % (data_onto_by_id[cat_id]['name'],
                                                      (len(fsids_to_move_oneclass) - count_moved_toLQ)))

        # sanity check for ALL classes
        if count_moved_toLQ < len(fsids_to_move_oneclass):
            print('------------------Something wrong when moving sounds from %s.' % (data_onto_by_id[cat_id]['name']))




    """ # apply STRONG filters to data_qual_sets********************************************************************"""

    # FILTER 1: Consider only leaf categories: 474 out of the initial 632
    # for o in data_qual_sets. o = catid
    # create a dict of dicts. The latter are key=o, and value the actual value (data_qual_sets[o])
    data_qual_sets_l = {o: data_qual_sets[o] for o in data_qual_sets if len(data_onto_by_id[o]['child_ids']) < 1}
    print 'Number of leaf categories: ' + str(len(data_qual_sets_l))
    # print()

    # plot
    nb_samples_cats_dev = compute_median(data_qual_sets_l)
    # plots before strong filters: all possible categories
    if FLAG_BOXPLOT:
        # boxplot number of examples per category in dev
        x_labels = 'estimated dev'
        fig_title = 'LEAVES: estimated number of clips in categories of DEV set'
        y_label = '# of audio clips'
        plot_boxplot(nb_samples_cats_dev, x_labels, fig_title, y_label)

    # remove some leaves manually
    # case A) there is a leaf that is accepted, but too specific. We prefer its parent
    # Pizzicato --> Violin
    # Alto Saxophone --> Saxo
    # case B) wind chime contains clips that should only be in Chime. Use Chime
    # In this order:
    leavesid_to_remove = ['/m/0d8_n', '/m/02pprs', '/m/026fgl']
    for catid in leavesid_to_remove:
        del data_qual_sets_l[catid]

    if FLAG_BARPLOT:
        # create variable with data for barplotting - function
        var_barplot = create_var_barplot(data_qual_sets_l, data_onto_by_id)

        # barplot with number of sounds per category
        # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
        idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
        # idx = np.argsort([data_onto_rich[tt]['nbHQ_dev'] for tt in range(len(data_onto_rich))])

        data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
        data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
        x_labels = list(var_barplot[val]['name'] for val in idx)
        y_label = '# of audio clips'
        fig_title = 'LEAVES: number of DEV (' + str(PERCENTAGE_DEV*100) +  '%) clips per category, split in HQdev and LQ, sorted by total'
        legenda = ('LQ', 'HQdev')
        plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)



    # FILTER 2: Apply duration filter: Within the 474 categories, keep sounds with durations [MINLEN: MAXLEN]
    # create copy for result of filter
    data_qual_sets_ld = copy.deepcopy(data_qual_sets_l)
    for catid, groups in data_qual_sets_ld.iteritems():
        data_qual_sets_ld[catid]['HQ'] = []
        data_qual_sets_ld[catid]['LQ'] = []

    for catid, groups in data_qual_sets_l.iteritems():
        for fsid in groups['HQ']:
            if (data_mapping[str(fsid)]['duration'] <= MAXLEN) and (data_mapping[str(fsid)]['duration'] >= MINLEN):
                data_qual_sets_ld[catid]['HQ'].append(fsid)

        for fsid in groups['LQ']:
            if (data_mapping[str(fsid)]['duration'] <= MAXLEN) and (data_mapping[str(fsid)]['duration'] >= MINLEN):
                data_qual_sets_ld[catid]['LQ'].append(fsid)


    # FILTER 2.1: NC license. Within the categories, discard sounds with NC licenses and sampling+
    # create copy for result of filter
    data_qual_sets_ldl = copy.deepcopy(data_qual_sets_ld)
    for catid, groups in data_qual_sets_ldl.iteritems():
        data_qual_sets_ldl[catid]['HQ'] = []
        data_qual_sets_ldl[catid]['LQ'] = []

    for catid, groups in data_qual_sets_ld.iteritems():
        for fsid in groups['HQ']:
            if data_mapping[str(fsid)]['license'].split('/')[-3] != 'by-nc' and\
                            data_mapping[str(fsid)]['license'].split('/')[-3] != 'sampling+':
                data_qual_sets_ldl[catid]['HQ'].append(fsid)

        for fsid in groups['LQ']:
            if data_mapping[str(fsid)]['license'].split('/')[-3] != 'by-nc' and\
                            data_mapping[str(fsid)]['license'].split('/')[-3] != 'sampling+':
                data_qual_sets_ldl[catid]['LQ'].append(fsid)


    # FILTER 3:  number of sounds with HQ>= MIN_HQ (what we proposed already)
    # o = catid. create a dict of dicts. the latter are just the dicts that fulfil the condition on MIN_HQ
    data_qual_sets_ld_HQ = {o: data_qual_sets_ldl[o] for o in data_qual_sets_ldl if
                            len(data_qual_sets_ldl[o]['HQ']) >= MIN_HQ}

    # data_qual_sets_ld_HQ means: every category with HQ and LQ after:
    # - taking leaves
    # - applying duration filter
    # - applying license filter (although not explicit in var name)
    # - num HQ > MIN_HQ

    print 'Number of leaf categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, duration [' + str(
        MINLEN) + ':' + str(MAXLEN) + '], and NC/sampling+ free: ' + str(len(data_qual_sets_ld_HQ))
    # print()

    # plot
    nb_samples_cats_dev = compute_median(data_qual_sets_ld_HQ)
    # plots before strong filters: all possible categories
    if FLAG_BOXPLOT:
        # boxplot number of examples per category in dev
        x_labels = 'estimated dev'
        fig_title = 'LEAVES | DUR | MIN_HQ: estimated number of clips in categories of DEV set'
        y_label = '# of audio clips'
        plot_boxplot(nb_samples_cats_dev, x_labels, fig_title, y_label)

    if FLAG_BARPLOT:
        # barplot with number of sounds per category
        # create variable with data for barplotting
        var_barplot = create_var_barplot(data_qual_sets_ld_HQ, data_onto_by_id)
        # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
        idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
        data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
        data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
        x_labels = list(var_barplot[val]['name'] for val in idx)
        y_label = '# of audio clips'
        fig_title = 'LEAVES | DUR | MIN_HQ: number of DEV ('\
                    + str(PERCENTAGE_DEV*100) +  '%) clips per category, split in HQdev and LQ, sorted by total'
        legenda = ('LQ', 'HQdev')
        plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)

        # barplot again sorted by HQdev
        idx = np.argsort([var_barplot[tt]['nbHQ_dev'] for tt in range(len(var_barplot))])
        data_up = list(var_barplot[val]['nbLQ'] for val in idx)
        data_bottom = list(var_barplot[val]['nbHQ_dev'] for val in idx)
        x_labels = list(var_barplot[val]['name'] for val in idx)
        y_label = '# of audio clips'
        fig_title = 'LEAVES | DUR | MIN_HQ: number of DEV ('\
                    + str(PERCENTAGE_DEV*100) + '%) clips per category, split in HQdev and LQ, sorted by HQdev'
        legenda = ('HQdev', 'LQ')
        plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)



    """ # apply FLEXIBLE filters to data_qual_sets*****************************************************************"""

    # print 'APPROACH ALFA is more strict: separate threshold on HQ and LQ'
    # # FILTER 4: MIN_LQ. number of sounds with LQ should not be less than MIN_LQ
    # data_qual_sets_ld_HQLQ = {o: data_qual_sets_ld_HQ[o] for o in data_qual_sets_ld_HQ if
    #                           len(data_qual_sets_ld_HQ[o]['LQ']) >= MIN_LQ}
    # print 'Number of leaf categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, and at least ' + str(
    #     MIN_LQ) + ' sounds with LQ labels, and of duration [' + str(MINLEN) + ':' + str(MAXLEN) + ']: ' + str(
    #     len(data_qual_sets_ld_HQLQ))
    #
    #
    # # FILTER 5: MIN_QE. to trust the LQ subset, we demand a minimum QE
    # data_qual_sets_ld_HQLQQE = {o: data_qual_sets_ld_HQLQ[o] for o in data_qual_sets_ld_HQLQ if
    #                             data_qual_sets_ld_HQLQ[o]['QE'] >= MIN_QE}
    # print 'Number of leaf categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, and at least ' + str(
    #     MIN_LQ) + ' sounds with LQ labels with a QE > ' + str(MIN_QE) + ', and of duration [' \
    #       + str(MINLEN) + ':' + str(MAXLEN) + ']: ' + str(len(data_qual_sets_ld_HQLQQE))
    #
    #
    # # plot
    # nb_samples_cats_dev = compute_median(data_qual_sets_ld_HQLQQE)
    # # plots before strong filters: all possible categories
    # if FLAG_BOXPLOT:
    #     # boxplot number of examples per category in dev
    #     x_labels = 'estimated dev'
    #     fig_title = 'ALFA - LEAVES | DUR | MIN_HQ_LQ_QE: estimated number of clips in categories of DEV set'
    #     y_label = '# of audio clips'
    #     plot_boxplot(nb_samples_cats_dev, x_labels, fig_title, y_label)
    #
    # if FLAG_BARPLOT:
    #     # barplot with number of sounds per category
    #     # create variable with data for barplotting
    #     var_barplot = create_var_barplot(data_qual_sets_ld_HQLQQE, data_onto_by_id)
    #     # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
    #     idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
    #     data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
    #     data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
    #     x_labels = list(var_barplot[val]['name'] for val in idx)
    #     y_label = '# of audio clips'
    #     fig_title = 'ALFA - LEAVES | DUR | MIN_HQ_LQ_QE: number of DEV ('\
    #                 + str(PERCENTAGE_DEV*100) +  '%) clips per category, split in HQdev and LQ, sorted by total'
    #     legenda = ('LQ', 'HQdev')
    #     plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)
    #
    # print 'Total Number of sounds in DEV set (estimated), there is a LOT of LQ: ' + str(sum(nb_samples_cats_dev))
    #
    #


    # print()
    print()
    print 'APPROACH BETA is less strict: HQdev + LQ > MIN_HQdev_LQ (joint)'
    # CASE BETA
    # FILTER 4: MIN_HQdev_LQ. number of sounds amounted between HQdev + LQ s>= MIN_HQdev_LQ
    # remember: LQprior is already contained in LQ (we just use it for prioritization within LQ and when QE is low)
    data_qual_sets_ld_HQLQb = {o: data_qual_sets_ld_HQ[o] for o in data_qual_sets_ld_HQ if
                               (len(data_qual_sets_ld_HQ[o]['LQ']) + np.floor(PERCENTAGE_DEV *len(data_qual_sets_ld_HQ[o]['HQ']))) >= MIN_HQdev_LQ}

    # sanity
    for catid, groups in data_qual_sets_ld_HQLQb.iteritems():
        if (len(groups['LQ']) + np.floor(PERCENTAGE_DEV *len(groups['HQ']))) < MIN_HQdev_LQ:
            print 'error in the category: ' + str(data_onto_by_id[str(catid)]['name'])

    print 'Number of leaf categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, and at least ' + str(
        MIN_HQdev_LQ) + ' sounds between HQdev and LQ labels, duration ['\
          + str(MINLEN) + ':' + str(MAXLEN) + '], and NC/sampling+ free: ' + str(len(data_qual_sets_ld_HQLQb))




    # FILTER 5: MIN_QE. to trust the LQ subset, we demand a minimum QE

    # old way: hard constraint on QE
    # data_qual_sets_ld_HQLQQEb = {o: data_qual_sets_ld_HQLQb[o] for o in data_qual_sets_ld_HQLQb if
    #                              data_qual_sets_ld_HQLQb[o]['QE'] >= MIN_QE}
    #
    # print 'Number of leaf categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, and at least ' + str(
    #     MIN_HQdev_LQ) + ' sounds between HQdev and LQ labels with a QE > ' + str(MIN_QE) + ', and of duration [' \
    #       + str(MINLEN) + ':' + str(MAXLEN) + ']: ' + str(len(data_qual_sets_ld_HQLQQEb))
    #

    # new way: consider that some categories may not meet MIN_QE, but they may have enough HQ to fill MIN_HQdev_LQ
    data_qual_sets_ld_HQLQQEb = {}
    for o in data_qual_sets_ld_HQLQb:
        if o == '/m/0bm02':
            b = 9

        if data_qual_sets_ld_HQLQb[o]['QE'] >= MIN_QE:
            # category with high QE
            # keep all data HQ, LQ, QE, update LQprior according to the filtering results
            data_qual_sets_ld_HQLQQEb[o] = data_qual_sets_ld_HQLQb[o]
            # make sure LQprior are in LQ (some were deleted due to duration and license filters)
            data_qual_sets_ld_HQLQQEb[o]['LQprior'] = \
               list(set(data_qual_sets_ld_HQLQb[o]['LQprior']) & set(data_qual_sets_ld_HQLQb[o]['LQ']))

        # from here on, cannot trust LQ

        # next two can be collapsed into one clause
        # category with enough HQ to fill all DEV and EVAL without anything of LQ
        elif np.floor(PERCENTAGE_DEV * len(data_qual_sets_ld_HQLQb[o]['HQ'])) >= MIN_HQdev_LQ:
            # keep HQ, put LQ = LQprior (that meet previous filters), keep LQprior and remove QE
            data_qual_sets_ld_HQLQQEb[o] = data_qual_sets_ld_HQLQb[o]
            data_qual_sets_ld_HQLQQEb[o]['LQ'] = \
                list(set(data_qual_sets_ld_HQLQb[o]['LQprior']) & set(data_qual_sets_ld_HQLQb[o]['LQ']))
            data_qual_sets_ld_HQLQQEb[o]['LQprior'] = \
                list(set(data_qual_sets_ld_HQLQb[o]['LQprior']) & set(data_qual_sets_ld_HQLQb[o]['LQ']))
            del data_qual_sets_ld_HQLQQEb[o]['QE']
            print '---category with enough HQ to fill DEV and EVAL without LQ: ' \
                  + str(data_onto_by_id[o]['name']) + ', with number of samples in HQ: ' \
                  + str(len(data_qual_sets_ld_HQLQb[o]['HQ']))
            print()

        # category with enough (HQ + LQprior) to fill all DEV and EVAL without the rest of LQ
        elif (np.floor(PERCENTAGE_DEV * len(data_qual_sets_ld_HQLQb[o]['HQ'])) + \
                len(list(set(data_qual_sets_ld_HQLQb[o]['LQprior']) & set(data_qual_sets_ld_HQLQb[o]['LQ'])))) >= MIN_HQdev_LQ:
            # keep HQ, put LQ = LQprior (that meet previous filters), keep LQprior and remove QE
            data_qual_sets_ld_HQLQQEb[o] = data_qual_sets_ld_HQLQb[o]
            data_qual_sets_ld_HQLQQEb[o]['LQ'] = \
                list(set(data_qual_sets_ld_HQLQb[o]['LQprior']) & set(data_qual_sets_ld_HQLQb[o]['LQ']))
            data_qual_sets_ld_HQLQQEb[o]['LQprior'] = \
                list(set(data_qual_sets_ld_HQLQb[o]['LQprior']) & set(data_qual_sets_ld_HQLQb[o]['LQ']))
            del data_qual_sets_ld_HQLQQEb[o]['QE']
            print '---category where only (HQ and LQprior) together, can fill DEV and EVAL: ' \
                  + str(data_onto_by_id[o]['name']) + ', with number of samples in HQ: ' \
                  + str(len(data_qual_sets_ld_HQLQb[o]['HQ'])) + ', and number of samples in LQprior: ' \
                  + str(len(data_qual_sets_ld_HQLQb[o]['LQprior']))
            print()

    # sanity check: LQprior must be a subset of LQ
    for catid, groups in data_qual_sets_ld_HQLQQEb.iteritems():
        if len(list(set(groups['LQprior']) & set(groups['LQ']))) != len(groups['LQprior']):
            sys.exit('mistake at LQprior filtering in leaves')

    print 'Number of leaf categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, at least ' + str(
        MIN_HQdev_LQ) + ' sounds between HQdev and LQ labels with a QE > ' + str(MIN_QE) + ', duration [' \
          + str(MINLEN) + ':' + str(MAXLEN) + '], and NC/sampling+ free: ' + str(len(data_qual_sets_ld_HQLQQEb))

    print 'List of selected LEAVES categories: \n'
    idx_print = 1
    for catid in data_qual_sets_ld_HQLQQEb.keys():
        print str(idx_print) + '-' + data_onto_by_id[str(catid)]['name']
        idx_print += 1

    # plot
    nb_samples_cats_dev = compute_median(data_qual_sets_ld_HQLQQEb)
    # plots before strong filters: all possible categories
    if FLAG_BOXPLOT:
        # boxplot number of examples per category in dev
        x_labels = 'estimated dev'
        fig_title = 'BETA - LEAVES | DUR | MIN_HQdev_LQ: estimated number of clips in categories of DEV set'
        y_label = '# of audio clips'
        plot_boxplot(nb_samples_cats_dev, x_labels, fig_title, y_label)

    if FLAG_BARPLOT:
        # barplot with number of sounds per category
        # create variable with data for barplotting
        var_barplot = create_var_barplot(data_qual_sets_ld_HQLQQEb, data_onto_by_id)
        # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
        idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
        data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
        data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
        x_labels = list(var_barplot[val]['name'] for val in idx)
        y_label = '# of audio clips'
        fig_title = 'BETA - LEAVES | DUR | MIN_HQdev_LQ_QE: number of DEV ('\
                    + str(PERCENTAGE_DEV*100) + '%) clips per category, split in HQdev and LQ, sorted by total'
        legenda = ('LQ', 'HQdev')
        plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)

        # barplot again sorted by HQdev
        idx = np.argsort([var_barplot[tt]['nbHQ_dev'] for tt in range(len(var_barplot))])
        data_up = list(var_barplot[val]['nbLQ'] for val in idx)
        data_bottom = list(var_barplot[val]['nbHQ_dev'] for val in idx)
        x_labels = list(var_barplot[val]['name'] for val in idx)
        y_label = '# of audio clips'
        fig_title = 'BETA - LEAVES | DUR | MIN_HQdev_LQ_QE: number of DEV ('\
                    + str(PERCENTAGE_DEV*100) + '%) clips per category, split in HQdev and LQ, sorted by HQdev'
        legenda = ('HQdev', 'LQ')
        plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)



    # print 'Total Number of sounds in DEV set (estimated), there is a LOT of LQ: ' + str(sum(nb_samples_cats_dev))
    # print()
    print '======================================================'
    print '\n\n\n'



    """ # from data_qual_sets to data_qual_sets_pparents  *****************************************************
    ***********************************************************************************************************
    ***********************************************************************************************************"""


    # So far we have considered only leafs. Now, there are parents such that all the childrens are discarded.
    # we could try to aggregate these children together (the part of them that it is ok)
    # with the parent (if applicable), and see if the resulting aggregated category meets the requirements.
    # how many categories are gained?

    # final set of valid leaf categories for dataset
    final_set_valid_leaves = [catid for catid in data_qual_sets_ld_HQLQQEb]

    # list all penultimate parents (parents at the penultimate level of the onto)
    penul_parents = []
    for catid, info in data_onto_by_id.iteritems():
        # if they have children
        if info['child_ids']:
            # but none of the children have further children
            flag_penul_parent = True
            for childid in info['child_ids']:
                if data_onto_by_id[str(childid)]['child_ids']:
                    flag_penul_parent = False
                    break
            if flag_penul_parent:
                penul_parents.append({'catid': catid, 'name': data_onto_by_id[str(catid)]['name']})
                # print (data_onto_by_id[str(catid)]['name'])

    print 'There are ' + str(len(penul_parents)) + ' penultimate parents\n'

    # how many of the penultimate parents have ALL children discarded for the dataset (due to ANY reason)?
    penul_parents_cand = []
    for penul_parent in penul_parents:
        flag_all_children_discarded = True
        for childid in data_onto_by_id[penul_parent['catid']]['child_ids']:
            if childid in final_set_valid_leaves:
                flag_all_children_discarded = False
                break
        if flag_all_children_discarded:
            penul_parents_cand.append({'catid': penul_parent['catid'], 'name': penul_parent['name']})
            # print (penul_parent['name'])

    print 'There are ' + str(len(penul_parents_cand)) + ' penultimate parents; ALL children either ' \
                                                        'not eligible (eg RED) or discarded for the dataset\n'

    # these penultimate parents with ALL children not eligible or discarded, are candidates to be part of the dataset
    # they are candidates to be "new children"

    # we have to populate them with their useless children (join them):
    # 1- only consider the children that do not have multiple parents.
    # For example, HISS is a child and has multiple parents (cat, steam, snake, onomatopeia)
    # u cannot say that a steam hiss is a cat hiss, so we cannot propagate from hiss to cat.
    # but the rest of the Cat's children can be populated (meouw, purr, etc)

    # 2-populate parents with valid children: HQ, LQ
    # concatenate lists of HQ and LQ (easy)? set
    # but the concepts of LQ and HQ are different when we populate higher in the hierarchy.
    # it can happen that in child it is LQ while it is HQ in parent: for example, if we have a meow and a purr,
    # it will be PNP on any of the children, but PP on the immediate parent CAT
    # we do not know this. it could also be that there is purr + speech (category outside the 'family'), and so
    # it must stay PNP both in children and parent.
    # The easiest thing, and more restrictive/demanding/ensuring better quality is:
    # join HQ and join LQ. whatever it is HQ in children, it should also be in parent.
    # and in this way, the LQ will be of better quality
    # but there are several cases to be consider. doing it exhaustively.

    # 3-recompute the QE, with votes... think cases: sound in children OR in father vs sound in both
    # could be contradictory votes. A cat purr candidate to meow




    # create data_qual_sets_pparents with keys with catids and empty dicts as values, filled with empty lists for HQ/LQ
    data_qual_sets_pparents = copy.deepcopy(data_qual_sets)
    for catid, groups in data_qual_sets_pparents.iteritems():
        data_qual_sets_pparents[catid].clear()
        data_qual_sets_pparents[catid]['HQ'] = []
        data_qual_sets_pparents[catid]['LQ'] = []
        data_qual_sets_pparents[catid]['LQprior'] = []

    penul_parents_cand_2filt = []
    counter_mult_parents = 0

    for penul_parent in penul_parents_cand:

        """ # 1) distribute children  *****************************************************
        ***********************************************************************************************************"""

        # if penul_parent['name'] == 'Saxophone':
        # if penul_parent['name'] == 'Chicken, rooster':
        # if penul_parent['name'] == 'Vehicle horn, car horn, honking':
        # if penul_parent['name'] == 'Thunderstorm':
        if penul_parent['name'] == 'Chime':
            a = 9

        count_weird_pop_fromHQ2LQ = 0
        children_valid_popul_QEhigh = []
        children_valid_popul_QElow = []
        for childid in data_onto_by_id[penul_parent['catid']]['child_ids']:

            # 1- for every child: if there is no multiple parents
            nb_parents = len([parentid for parentid in data_onto_by_id if childid in data_onto_by_id[str(parentid)]['child_ids']])

            if nb_parents == 1:
                # only consider children withOUT multiple parents

                # checking minimum QE for every category individually, for simplicity
                if data_qual_sets[str(childid)]['QE'] > MIN_QE:
                    # the number of votes > MIN_VOTES_CAT was checked before. if not, QE = 0 already
                    # the QE requirement is met. Leverage both HQ and LQ from the child
                    children_valid_popul_QEhigh.append(childid)
                elif data_qual_sets[str(childid)]['HQ']:
                    # the QE requirement is NOT met. Leverage only HQ and LQprior from the child
                    children_valid_popul_QElow.append(childid)

            elif nb_parents == 0:
                sys.exit('mistake at multiple parents computation')
            else:
                counter_mult_parents += 1

        # here we have children_valid_popul_QEhigh with the valid children for penul_parent, including ['QE'] > MIN_QE:
        # and children_valid_popul_QElow with the valid children for penul_parent that do not meet QE


        # 2 - when populating there are several cases possible
        # a) sound appears twice in HQ, or on LQ. Do set to keep it once. sanity check.

        # b,c,d are only the most typical of a series of combinations that may happen. covered below in basic if-clauses

        # b) sound appears in parent_HQ and child_LQ:
        # if we have a meow and a purr (PNP) or if you are Unsure between a purr/meow
        # count how many times this happens. leave in parent_HQ

        # c) sound appears in parent_LQ and child_HQ: is this possible? in theory it should not:
        # - absence of noise in one means it is also absent in the other
        # - PP in children, means PP in parent.
        # monitor this. it could happen due to different subjective evaluation.
        # same file rated differently by different people.
        # count how many times this happens. leave in parent_LQ.

        # d) once you merge the children, it can happen that a sound is in HQ and LQ
        # sound with a meow and a purr (with both tags) can be candidate in both siblings
        # it can be validated as PP by userA (wrong doing), goes to HQ, but
        # but as PNP by userB in purr, goes to LQ
        # goes to parent_HQ

        # extend lists of HQ/LQ sounds of all children to populate
        children_id_2_parent_HQ = []
        children_id_2_parent_LQ = []
        # these will be a list of ASO ids, ie categories from where to pipulate data
        # the extension is done at the level of data_qual_sets, ie NO PROCESSING, LQprior is just PNP group
        # we are implmenting the QE filters here. Only the suitable data is accepted, based on QE

        if children_valid_popul_QElow and not children_valid_popul_QEhigh:
            # the QE requirement is NOT met never. Leverage only HQ and LQprior from the children
            # ONLY children's LQprior will be sent to parent (to LQprior and LQ)
            children_id_2_parent_HQ.extend(children_valid_popul_QElow)

        elif not children_valid_popul_QElow and children_valid_popul_QEhigh:
            # the QE requirement is met in all cases. Leverage both HQ and LQ from all the children
            # children's LQprior will be sent to parent (only to LQprior to prioritize LQ content)
            children_id_2_parent_HQ.extend(children_valid_popul_QEhigh)

            # to count for acceptance: children's LQ will be sent to LQ of parent
            children_id_2_parent_LQ.extend(children_valid_popul_QEhigh)

        elif children_valid_popul_QElow and children_valid_popul_QEhigh:
            # some children did not meet the QE req while others did meet it.

            # some children did not meet the QE req. Leverage only HQ and LQprior from them
            # ONLY children's LQprior will be sent to parent (to LQprior and LQ)
            children_id_2_parent_HQ.extend(children_valid_popul_QElow)

            # Others met the QE req.  Leverage both HQ and LQ from all them
            # children's LQprior will be sent to parent (only to LQprior to prioritize LQ content)
            children_id_2_parent_HQ.extend(children_valid_popul_QEhigh)

            # to count for acceptance: children's LQ will be sent to LQ of parent
            children_id_2_parent_LQ.extend(children_valid_popul_QEhigh)

        # else:
            # print 'no children to populate'

        if children_id_2_parent_HQ:
            # at least we have to populate HQ and LQprior
            # consider HQ
            # grab all valid children, merge into list, set
            children_joint_lists_HQ = [data_qual_sets[str(childid)]['HQ'] for childid in children_id_2_parent_HQ]
            children_joint_HQ = list(set(list(itertools.chain.from_iterable(children_joint_lists_HQ))))
            # the latter should not produce duplicates

            # distribution of children------------------------------------make function
            for id in children_joint_HQ:
                # several options
                # sound only in children
                if id not in data_qual_sets[penul_parent['catid']]['HQ'] and id not in data_qual_sets[penul_parent['catid']]['LQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['HQ'].append(id)

                    # sound also in parent_HQ
                elif id in data_qual_sets[penul_parent['catid']]['HQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['HQ'].append(id)

                    # sound also in parent_LQ
                elif id in data_qual_sets[penul_parent['catid']]['LQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['LQ'].append(id)
                    count_weird_pop_fromHQ2LQ += 1

            # LQprior is always sent both to LQprior (prior) and LQ (acceptance) of the parent
            # if the child is QEhigh, LQ will be copied too, so LQprior will be duplicated in LQ (need set)
            # if the child is QElow, LQ will not be copied (LQprior will be unique in LQ)
            children_joint_lists_LQprior= \
                [data_qual_sets[str(childid)]['LQprior'] for childid in children_id_2_parent_HQ]
            children_joint_LQprior = list(set(list(itertools.chain.from_iterable(children_joint_lists_LQprior))))
            # the latter could produce duplicates

            for id in children_joint_LQprior:
                # children_joint_LQprior is list of ids of LQ sounds with priority (basically the PNP group)
                # same philosophy as LQ-if
                # sound only in children. several options
                if id not in data_qual_sets[penul_parent['catid']]['HQ'] and \
                        id not in data_qual_sets[penul_parent['catid']]['LQ']:
                    if id in children_joint_LQprior and id not in children_joint_HQ:
                        # normal case
                        data_qual_sets_pparents[penul_parent['catid']]['LQprior'].append(id)
                        data_qual_sets_pparents[penul_parent['catid']]['LQ'].append(id)
                    elif id in children_joint_LQprior and id in children_joint_HQ:
                        # priority for one child and HQ for a different child. unusual
                        data_qual_sets_pparents[penul_parent['catid']]['HQ'].append(id)

                    # sound also in parent_HQ
                elif id in data_qual_sets[penul_parent['catid']]['HQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['HQ'].append(id)

                    # sound also in parent_LQ
                elif id in data_qual_sets[penul_parent['catid']]['LQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['LQprior'].append(id)
                    data_qual_sets_pparents[penul_parent['catid']]['LQ'].append(id)

        if children_id_2_parent_LQ:
            # we also have to populate LQ (this only happens if children_joint_HQ exists too)
            children_joint_lists_LQ = [data_qual_sets[str(childid)]['LQ'] for childid in children_id_2_parent_LQ]
            children_joint_LQ = list(set(list(itertools.chain.from_iterable(children_joint_lists_LQ))))
            # the latter could produce duplicates

            for id in children_joint_LQ:
                # sound only in children. several options
                if id not in data_qual_sets[penul_parent['catid']]['HQ'] and \
                        id not in data_qual_sets[penul_parent['catid']]['LQ']:
                    if id in children_joint_LQ and id not in children_joint_HQ:
                        data_qual_sets_pparents[penul_parent['catid']]['LQ'].append(id)
                    elif id in children_joint_LQ and id in children_joint_HQ:
                        data_qual_sets_pparents[penul_parent['catid']]['HQ'].append(id)

                    # sound also in parent_HQ
                elif id in data_qual_sets[penul_parent['catid']]['HQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['HQ'].append(id)

                    # sound also in parent_LQ
                elif id in data_qual_sets[penul_parent['catid']]['LQ']:
                    data_qual_sets_pparents[penul_parent['catid']]['LQ'].append(id)

        # so far: the children have been distributed in a disjoint fashion

        """ # 2) distribute parent  *****************************************************
        ***********************************************************************************************************"""

        # now, independent of the children, add the parents if QE allows it
        # these categories were not considered before because they were not leafs
        FLAG_PARENT_IN = 0
        if data_qual_sets[penul_parent['catid']]['QE'] >= MIN_QE:
            # add both HQ and LQ
            FLAG_PARENT_IN = 2
            data_qual_sets_pparents[penul_parent['catid']]['HQ'].extend(data_qual_sets[penul_parent['catid']]['HQ'])
            data_qual_sets_pparents[penul_parent['catid']]['LQ'].extend(data_qual_sets[penul_parent['catid']]['LQ'])
            # add LQprior ONLY to LQprior for prioritizing
            data_qual_sets_pparents[penul_parent['catid']]['LQprior'].\
                extend(data_qual_sets[penul_parent['catid']]['LQprior'])

        else:
            # we cannot trust LQ.
            FLAG_PARENT_IN = 1
            # Add HQ
            data_qual_sets_pparents[penul_parent['catid']]['HQ'].extend(data_qual_sets[penul_parent['catid']]['HQ'])
            # add LQprior both to LQprior (prior) and LQ (acceptance)
            data_qual_sets_pparents[penul_parent['catid']]['LQprior']. \
                extend(data_qual_sets[penul_parent['catid']]['LQprior'])
            data_qual_sets_pparents[penul_parent['catid']]['LQ'].\
                extend(data_qual_sets[penul_parent['catid']]['LQprior'])

        # remove duplicates from the process
        data_qual_sets_pparents[penul_parent['catid']]['HQ'] = \
            list(set(data_qual_sets_pparents[penul_parent['catid']]['HQ']))
        data_qual_sets_pparents[penul_parent['catid']]['LQ'] = \
            list(set(data_qual_sets_pparents[penul_parent['catid']]['LQ']))
        data_qual_sets_pparents[penul_parent['catid']]['LQprior'] = \
            list(set(data_qual_sets_pparents[penul_parent['catid']]['LQprior']))

        """ # 3) sanity checks  *****************************************************
        ***********************************************************************************************************"""

        # sanity check: LQprior must be a subset of LQ
        if len(list(set(data_qual_sets_pparents[penul_parent['catid']]['LQprior']) &
                            set(data_qual_sets_pparents[penul_parent['catid']]['LQ']))) !=\
                len(data_qual_sets_pparents[penul_parent['catid']]['LQprior']):
            sys.exit('mistake at LQprior distribution in penultimate parents')

        # sanity check: number of sounds must be equal before and after population
        # sounds after the processing
        nb_sounds_postpop = len(data_qual_sets_pparents[penul_parent['catid']]['HQ']) + \
                            len(data_qual_sets_pparents[penul_parent['catid']]['LQ'])

        if children_valid_popul_QElow and not children_valid_popul_QEhigh:
            # only children with low QE. count only HQ and LQprior

            # concatenating lists and set
            if FLAG_PARENT_IN == 2:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQprior +
                                           data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQ']))
            elif FLAG_PARENT_IN == 1:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQprior +
                                           data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQprior']))
            elif FLAG_PARENT_IN == 0:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQprior))
            else:
                sys.exit('sanity check number of sounds')

            if nb_sounds_prepop != nb_sounds_postpop:
                print(catid)
                sys.exit('number of sounds is not equal before and after population - children_valid_popul_QElow')

        elif not children_valid_popul_QElow and children_valid_popul_QEhigh:
            # only children with high QE. count all: HQ and LQ
            # LQprior is already counted in LQ

            # concatenating lists and set
            if FLAG_PARENT_IN == 2:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQ +
                                           data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQ']))
            elif FLAG_PARENT_IN == 1:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQ +
                                           data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQprior']))
            elif FLAG_PARENT_IN == 0:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQ))
            else:
                sys.exit('sanity check number of sounds')

            if nb_sounds_prepop != nb_sounds_postpop:
                print(penul_parent)
                sys.exit('number of sounds is not equal before and after population - children_valid_popul_QEhigh')

        elif children_valid_popul_QElow and children_valid_popul_QEhigh:
            # both tyoes of children mixed:
            # some with high QE. count all: HQ and LQ (LQprior is already counted in LQ)
            # some with low QE. count only HQ and LQprior

            # concatenating lists and set
            if FLAG_PARENT_IN == 2:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQ + children_joint_LQprior +
                                           data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQ']))
            elif FLAG_PARENT_IN == 1:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQ + children_joint_LQprior +
                                           data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQprior']))
            elif FLAG_PARENT_IN == 0:
                nb_sounds_prepop = len(set(children_joint_HQ + children_joint_LQ + children_joint_LQprior))
            else:
                sys.exit('sanity check number of sounds')

            if nb_sounds_prepop != nb_sounds_postpop:
                print(penul_parent)
                sys.exit('number of sounds is not equal before and after population - children_valid_popul_QEhigh')

        else:
            # no population from the children
            # concatenating lists and set
            if FLAG_PARENT_IN == 2:
                nb_sounds_prepop = len(set(data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQ']))
            elif FLAG_PARENT_IN == 1:
                nb_sounds_prepop = len(set(data_qual_sets[penul_parent['catid']]['HQ'] +
                                           data_qual_sets[penul_parent['catid']]['LQprior']))
            elif FLAG_PARENT_IN == 0:
                nb_sounds_prepop = 0
            else:
                sys.exit('sanity check number of sounds')

            if nb_sounds_prepop != nb_sounds_postpop:
                print(penul_parent)
                sys.exit('number of sounds is not equal before and after population - no population from children')

        # how many penultimate parents (populated or not) are we considering?)
        # sometimes we populate them, having parent + child(ren)
        # or we dont consider the children but we consider the parent (for the first time)
        # or we dont consider the parent (due to QE) but the aggregation of the children
        if data_qual_sets_pparents[penul_parent['catid']]['HQ']:
            # if we actually have a new candidate category, sanity check

            # HQ and LQ must be disjoint groups
            if list(set(data_qual_sets_pparents[penul_parent['catid']]['HQ']) &
                    set(data_qual_sets_pparents[penul_parent['catid']]['LQ'])):
                print(penul_parent)
                sys.exit('data_qual_sets_pparents has not disjoint groups. beware population')

            penul_parents_cand_2filt.append(penul_parent)

    print 'There are ' + str(len(penul_parents_cand_2filt)) + ' penultimate parents entering the filtering stage, ie ' \
                                                              'after MultParents, indiv QE filter and population\n'


# remove empty categories from data_qual_sets_pparents (keep only the penultimate parents)
data_qual_sets_pparents_clean = {o: data_qual_sets_pparents[o] for o in data_qual_sets_pparents if data_qual_sets_pparents[o]['HQ']}
print 'Number of added penultimate parents entering the filtering stage: ' + str(len(data_qual_sets_pparents_clean))

# sanity check: dict of valid children and parents are disjoint at this point
if any(k in data_qual_sets_ld_HQLQQEb for k in data_qual_sets_pparents_clean):
    sys.exit('dicts of parent and children are not disjoint. DANGER')



""" # 4) once we have the new children, repeat filtering as if they were real children. ***********************
***********************************************************************************************************"""
# apply filters on LQ, then intersection with LQprior
# FILTER 2: Apply duration filter: Within the categories, keep sounds with durations [MINLEN: MAXLEN]
# create copy for next filter
data_qual_sets_pparents_d = copy.deepcopy(data_qual_sets_pparents_clean)
for catid, groups in data_qual_sets_pparents_d.iteritems():
    data_qual_sets_pparents_d[catid]['HQ'] = []
    data_qual_sets_pparents_d[catid]['LQ'] = []

for catid, groups in data_qual_sets_pparents_clean.iteritems():
    for fsid in groups['HQ']:
        if (data_mapping[str(fsid)]['duration'] <= MAXLEN) and (data_mapping[str(fsid)]['duration'] >= MINLEN):
            data_qual_sets_pparents_d[catid]['HQ'].append(fsid)

    for fsid in groups['LQ']:
        if (data_mapping[str(fsid)]['duration'] <= MAXLEN) and (data_mapping[str(fsid)]['duration'] >= MINLEN):
            data_qual_sets_pparents_d[catid]['LQ'].append(fsid)


# FILTER 2.1: NC license. Within the categories, discard sounds with NC licenses and sampling+
# create copy for result of filter
data_qual_sets_pparents_dl = copy.deepcopy(data_qual_sets_pparents_d)
for catid, groups in data_qual_sets_pparents_dl.iteritems():
    data_qual_sets_pparents_dl[catid]['HQ'] = []
    data_qual_sets_pparents_dl[catid]['LQ'] = []

for catid, groups in data_qual_sets_pparents_d.iteritems():
    for fsid in groups['HQ']:
        if data_mapping[str(fsid)]['license'].split('/')[-3] != 'by-nc' and \
                        data_mapping[str(fsid)]['license'].split('/')[-3] != 'sampling+':
            data_qual_sets_pparents_dl[catid]['HQ'].append(fsid)

    for fsid in groups['LQ']:
        if data_mapping[str(fsid)]['license'].split('/')[-3] != 'by-nc' and \
                        data_mapping[str(fsid)]['license'].split('/')[-3] != 'sampling+':
            data_qual_sets_pparents_dl[catid]['LQ'].append(fsid)


# FILTER 3: number of sounds with HQ>= MIN_HQ
# o = catid. create a dict of dicts. the latter are just the dicts that fulfil the condition on MIN_HQ
data_qual_sets_pparents_d_HQ = {o: data_qual_sets_pparents_dl[o] for o in data_qual_sets_pparents_dl if
                        len(data_qual_sets_pparents_dl[o]['HQ']) >= MIN_HQ}

# data_qual_sets_pparents_d_HQ means: every category with HQ and LQ after:
# - applying duration filter
# - applying license filter (although not explicit in var name)

print 'Number of pop parent categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, duration [' + \
      str(MINLEN) + ':' + str(MAXLEN) + '], and NC/sampling+ free: ' + str(len(data_qual_sets_pparents_d_HQ))
# print()


print()
print()
print 'APPROACH BETA is less strict: HQdev + LQ > MIN_HQdev_LQ (joint)'
# CASE BETA
# FILTER 4: MIN_HQdev_LQ. number of sounds amounted between HQdev + LQ s>= MIN_HQdev_LQ
# LQprior is already included in LQ
data_qual_sets_pparents_d_HQLQb = {o: data_qual_sets_pparents_d_HQ[o] for o in data_qual_sets_pparents_d_HQ if
                           (len(data_qual_sets_pparents_d_HQ[o]['LQ']) + np.floor(
                               PERCENTAGE_DEV * len(data_qual_sets_pparents_d_HQ[o]['HQ']))) >= MIN_HQdev_LQ}

# make sure LQprior are within LQ (some were deleted due to duration and license filters)
for o in data_qual_sets_pparents_d_HQLQb:
    data_qual_sets_pparents_d_HQLQb[o]['LQprior'] = \
        list(set(data_qual_sets_pparents_d_HQLQb[o]['LQprior']) & set(data_qual_sets_pparents_d_HQLQb[o]['LQ']))

# no need to test for QE of the categories because it is already done.
# the LQ of children and parent is only considered if its QE is higher than threshold, individually

# sanity check
for catid, groups in data_qual_sets_pparents_d_HQLQb.iteritems():
    if (len(groups['LQ']) + np.floor(PERCENTAGE_DEV * len(groups['HQ']))) < MIN_HQdev_LQ:
        print 'error at the category: ' + str(data_onto_by_id[str(catid)]['name'])

print 'Number of pop parent categories with at least ' + str(MIN_HQ) + ' sounds with HQ labels, at least ' + str(
    MIN_HQdev_LQ) + ' sounds between HQdev and LQ labels, duration [' \
      + str(MINLEN) + ':' + str(MAXLEN) + '], and NC/sampling+ free: ' + str(len(data_qual_sets_pparents_d_HQLQb))

if FLAG_BARPLOT_PARENT:
    # barplot with number of sounds per category
    # create variable with data for barplotting
    var_barplot = create_var_barplot(data_qual_sets_pparents_d_HQLQb, data_onto_by_id)
    # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
    idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
    data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
    data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
    x_labels = list(var_barplot[val]['name'] for val in idx)
    y_label = '# of audio clips'
    fig_title = 'BETA - PPARENTS | DUR | MIN_HQdev_LQ: number of DEV (' \
                + str(PERCENTAGE_DEV * 100) + '%) clips per category, split in HQdev and LQ, sorted by total'
    legenda = ('LQ', 'HQdev')
    plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)

    # barplot again sorted by HQdev
    idx = np.argsort([var_barplot[tt]['nbHQ_dev'] for tt in range(len(var_barplot))])
    data_up = list(var_barplot[val]['nbLQ'] for val in idx)
    data_bottom = list(var_barplot[val]['nbHQ_dev'] for val in idx)
    x_labels = list(var_barplot[val]['name'] for val in idx)
    y_label = '# of audio clips'
    fig_title = 'BETA - PPARENTS | DUR | MIN_HQdev_LQ: number of DEV (' \
                + str(PERCENTAGE_DEV * 100) + '%) clips per category, split in HQdev and LQ, sorted by HQdev'
    legenda = ('HQdev', 'LQ')
    plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)




""" 5) Merge both dictionaries and save********************************************************************************
****************************************************************************************************************"""

# joint both dicts
dataset_final_prepro = dict(data_qual_sets_ld_HQLQQEb)  # or orig.copy()
dataset_final_prepro.update(data_qual_sets_pparents_d_HQLQb)

# FINAL SANITY CHECKS for whole dataset:
for catid, groups in dataset_final_prepro.iteritems():
    # there must be no NC/sampling+ license (we will provide a license file with the dataset)
    for groupid, group in groups.iteritems():
        if groupid != 'QE':
            for fsid in group:
                if data_mapping[str(fsid)]['license'].split('/')[-3] == 'by-nc':
                    sys.exit('There are NC sounds. FATAL ERROR')
                elif data_mapping[str(fsid)]['license'].split('/')[-3] == 'sampling+':
                    sys.exit('There are sampling+ sounds. FATAL ERROR')

    # HQ and LQ groups in dataset_final_prepro should be disjoint
    if list(set(dataset_final_prepro[catid]['HQ']) & set(dataset_final_prepro[catid]['LQ'])):
        # print('\n something unexpetected happened in the mapping********************* \n')
        print(catid)
        sys.exit('dataset_final_prepro has not disjoint groups')

    # LQprior must be a subset of LQ
    if len(list(set(groups['LQprior']) & set(groups['LQ']))) != len(groups['LQprior']):
        sys.exit('dataset_final_prepro: mistake at LQprior filtering in leaves')


print 'Number of final categories: ' + str(len(dataset_final_prepro))
print 'List of selected categories out of PRE-PRO: \n'

idx_print = 1
for catid, groups in dataset_final_prepro.iteritems():
    # print str(idx_print) + '-' + data_onto_by_id[str(catid)]['name']
    print "%-2d-%-25s:  HQ: %-3d - LQ: %-5d - LQprior: %-3d" % (idx_print, data_onto_by_id[str(catid)]['name'],
                                                      len(groups['HQ']), len(groups['LQ']), len(groups['LQprior']))
    idx_print += 1
# %-12i
if FLAG_BARPLOT_PARENT:
    # barplot with number of sounds per category
    # create variable with data for barplotting
    var_barplot = create_var_barplot(dataset_final_prepro, data_onto_by_id)
    # sort by ascending number of TRAINING sounds in category (70% HQ + LQ)
    idx = np.argsort([var_barplot[tt]['nbtotal_tr'] for tt in range(len(var_barplot))])
    data_bottom = list(var_barplot[val]['nbLQ'] for val in idx)
    data_up = list(var_barplot[val]['nbHQ_dev'] for val in idx)
    x_labels = list(var_barplot[val]['name'] for val in idx)
    y_label = '# of audio clips'
    fig_title = 'BETA - PPARENTS | DUR | MIN_HQdev_LQ: number of DEV (' \
                + str(PERCENTAGE_DEV * 100) + '%) clips per category, split in HQdev and LQ, sorted by total'
    legenda = ('LQ', 'HQdev')
    plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)

    # barplot again sorted by HQdev
    idx = np.argsort([var_barplot[tt]['nbHQ_dev'] for tt in range(len(var_barplot))])
    data_up = list(var_barplot[val]['nbLQ'] for val in idx)
    data_bottom = list(var_barplot[val]['nbHQ_dev'] for val in idx)
    x_labels = list(var_barplot[val]['name'] for val in idx)
    y_label = '# of audio clips'
    fig_title = 'BETA - PPARENTS | DUR | MIN_HQdev_LQ: number of DEV (' \
                + str(PERCENTAGE_DEV * 100) + '%) clips per category, split in HQdev and LQ, sorted by HQdev'
    legenda = ('HQdev', 'LQ')
    plot_barplot(data_bottom, data_up, x_labels, y_label, fig_title, legenda)

# save
# with open('dataset_final_prepro.json', 'w') as fp:
#     json.dump(dataset_final_prepro, fp)



"""**********************************END OF INITIAL STAGE: *****************************************************
****************************************************************************************************************"""




"""**********************************BEGIN POST-PROCESSING STAGE: *****************************************************
****************************************************************************************************************"""

# ---------------------- REMOVE CATEGORIES ---------------------- #
category_id_to_remove = ['/m/0c1dj', '/m/07phxs1', '/m/02rr_', '/m/07s0s5r', 
                         '/m/0l14qv', '/m/05jcn', '/m/025l19', '/m/01b9nn', 
                         '/m/01jnbd', '/m/05mxj0q', '/m/06mb1', '/m/02hnl', 
                         '/m/02zsn', '/m/07r660_', '/t/dd00093', '/m/01vfsf', '/m/0912c9']
map(dataset_final_prepro.pop, set(category_id_to_remove) & set(dataset_final_prepro.keys()))


# --------------------------------------------------------------- #

# --------------- REMOVE MULTIPLE LABELED SOUNDS ---------------- #
all_sound_ids = []
for _, category_sets in dataset_final_prepro.iteritems():
    for set_name in ['HQ', 'LQ']:
        all_sound_ids += category_sets[set_name]

sound_to_remove = [s for s in all_sound_ids if all_sound_ids.count(s)>1]

# add duplicate in Freesound
sound_to_remove += [325477, 235363] # similar to 325480 and 235962 respectivly

for _, category_sets in dataset_final_prepro.iteritems():
    for set_name in ['HQ', 'LQ', 'LQprior']:
        for s in sound_to_remove:
            if s in category_sets[set_name]:
                category_sets[set_name].remove(s)

# --------------------------------------------------------------- #

# -------------------------------------------------------------------------- #
# ------------------------ CREATE SPLIT, FILES, ... ------------------------ #
# -------------------------------------------------------------------------- #
print '\n\n'
#dataset_final_prepro = json.load(open('', 'rb'))
result_final_HQ = {node_id: value['HQ'] for node_id, value in dataset_final_prepro.iteritems()}


# ---------------------- EXTRACT HQ SOUNDS --------------------- #
all_ids_HQ = []
for key, value in result_final_HQ.iteritems():
        all_ids_HQ += value
all_ids_HQ = set(all_ids_HQ)

sounds_with_labels_HQ = {sound_id:[] for sound_id in all_ids_HQ}
for node_id in result_final_HQ:
    for s in result_final_HQ[node_id]:
        sounds_with_labels_HQ[s].append(node_id)

# how many sounds with multiple labels
print 'Total amount of HQ sounds with more than one label: {0} samples'.format(
    len([1 for sound_id in sounds_with_labels_HQ if 
         len(sounds_with_labels_HQ[sound_id])>1]))

print 'Total Number of HQ labeled sounds: {}'.format(len(sounds_with_labels_HQ))


# --------------------------------------------------------------- #

# --------------------- SPLIT DEV EVAL HQ ------------------------#
print '\n SPLIT HQ \n'

# STRUCTURE DATA, SPLIT SINGLE/MULTIPLE LABELED SOUNDS
sounds_single = {s:sounds_with_labels_HQ[s] for s in sounds_with_labels_HQ if len(sounds_with_labels_HQ[s])==1}
sounds_multiple = {s:sounds_with_labels_HQ[s] for s in sounds_with_labels_HQ if len(sounds_with_labels_HQ[s])>1}

data_single = {r:[] for r in result_final_HQ}
for s in sounds_single:
    data_single[sounds_single[s][0]].append(s)

print 'Number of single labeled HQ sounds: {0}'.format(len(sounds_single))
    
data_multiple = {r:[] for r in result_final_HQ}
for s in sounds_multiple:
    for i in sounds_multiple[s]: 
        data_multiple[i].append(s)

print 'Number of multi-labeled HQ sounds: {0}'.format(len(sounds_multiple))

# ORDER BY DURATION
data_single_dur = {r:sorted([(s, data_mapping[str(s)]['duration']) for s in data_single[r]], key=lambda c:c[1]) for r in data_single}
data_multiple_dur = {r:sorted([(s, data_mapping[str(s)]['duration']) for s in data_multiple[r]], key=lambda c:c[1]) for r in data_multiple}

# SPLIT DEV/EVAL FOR SINGLE LABELED WITH RATIO 7:3 BASED ON DURATION
rule32 = ['dev', 'eval', 'dev', 'eval', 'dev']
rule73 = ['dev', 'eval', 'dev', 'dev', 'eval', 'dev', 'dev', 'eval', 'dev', 'dev']
data_single_dev = {r:[] for r in data_single_dur}
data_single_eval = {r:[] for r in data_single_dur}
for r in data_single_dur:
    for idx, s in enumerate(data_single_dur[r]):
        if rule73[idx%len(rule73)] == 'dev':
            data_single_dev[r].append(s[0])
        elif rule73[idx%len(rule73)] == 'eval':
            data_single_eval[r].append(s[0])
            
## RANDOMLY ADDING MULTIPLE LABELED WITH RATIO 7:3
data_dev_HQ = data_single_dev
data_eval = data_single_eval
for idx, s in enumerate(sounds_multiple):
    if rule73[idx%len(rule73)] == 'dev':
        for node_id in sounds_multiple[s]:
            data_dev_HQ[node_id].append(s)
    elif rule73[idx%len(rule73)] == 'eval':
        for node_id in sounds_multiple[s]:
            data_eval[node_id].append(s)

# EXPORT DATASET
ontology_by_id = {o['id']:o for o in data_onto}
dataset_dev_HQ = [{'name': ontology_by_id[node_id]['name'], 
                'audioset_id': node_id,
                'sound_ids': data_dev_HQ[node_id],
               } for node_id in data_dev_HQ]
dataset_eval = [{'name': ontology_by_id[node_id]['name'], 
                'audioset_id': node_id,
                'sound_ids': data_eval[node_id],
               } for node_id in data_eval]

        
# --------------------------------------------------------------- #

# --------------------- ADD LQ TO DEV SET ------------------------#
print '\n ADD LQ TO DEV SET'

data_dev = copy.deepcopy(data_dev_HQ)
#for node_id in data_dev.keys():
#    data_dev[node_id] += dataset_final_prepro[node_id]['LQ']
data_dev_LQ = {node_id: value['LQ'] for node_id, value in dataset_final_prepro.iteritems()}

for node_id, value in dataset_final_prepro.iteritems():
    if 'LQprior' not in value:
        value['LQprior'] = []
data_dev_LQpior = {node_id: value['LQprior'] for node_id, value in dataset_final_prepro.iteritems()}

dataset_dev_LQ = [{'name': ontology_by_id[node_id]['name'], 
                'audioset_id': node_id,
                'sound_ids': dataset_final_prepro[node_id]['LQ'],
               } for node_id in dataset_final_prepro]


# ---------------------------------------------------------------- #

# ------------------------ SELECT LQ SET ------------------------- #
MAX_NUM_SOUND_DEV = 300
selected_LQ = {node_id: [] for node_id in dataset_final_prepro}  # store the selected LQ prior and LQ

# ADD FIRST LQprior
for node_id in data_dev.keys():
    num_to_add = min(MAX_NUM_SOUND_DEV - len(data_dev[node_id]), len(data_dev_LQpior[node_id]))
    data_dev[node_id] += data_dev_LQpior[node_id][:num_to_add]
    selected_LQ[node_id] += data_dev_LQpior[node_id][:num_to_add]
    
# FILTER OUT LQprior from LQ
data_dev_LQ_wo_prior = {}
for node_id in data_dev_LQ:
    data_dev_LQ_wo_prior[node_id] = list(set(data_dev_LQ[node_id])-set(data_dev_LQpior[node_id]))

# ORDER BY NUM DOWNLOADS
for node_id in data_dev_LQ_wo_prior.keys():
    ll = []
    for fs_id in data_dev_LQ_wo_prior[node_id]:
        try:
            if data_mapping[str(fs_id)]['num_downloads']:
                ll.append((fs_id, data_mapping[str(fs_id)]['num_downloads']))
            else:
                ll.append((fs_id, 0))
        except:
            ll.append((fs_id, 0))
    freesound_ids_with_num_downloads = sorted(ll, key=lambda x: x[1], reverse=True)
    data_dev_LQ_wo_prior[node_id] = [fs_id_num_downloads[0] for fs_id_num_downloads in freesound_ids_with_num_downloads]

# ADD LQ TO DEV SET UNTIL REACHING MAX 300 SOUNDS
for node_id in data_dev.keys():
    num_to_add = min(MAX_NUM_SOUND_DEV - len(data_dev[node_id]), len(data_dev_LQ_wo_prior[node_id]))
    data_dev[node_id] += data_dev_LQ_wo_prior[node_id][:num_to_add]
    selected_LQ[node_id] += data_dev_LQ_wo_prior[node_id][:num_to_add]
    
dataset_dev = [{'name': ontology_by_id[node_id]['name'], 
                'audioset_id': node_id,
                'sound_ids': data_dev[node_id],
               } for node_id in data_dev]


""""***********************************************************************************************************"""
""""PACK EFFECT"""
""""***********************************************************************************************************"""
# # starting point is:
# # selected_LQ: dict with ALL the LQ sounds selected for the dataset (in some categories there were much more LQ
# # result_final_HQ: dict with ALL the HQ sounds of the dataset
#
# client = freesound.FreesoundClient()
# client.set_token("eaa4f46407adf86c35c5d5796fd6ea8b05515dca", "token")
#
# pack_status_HQ_per_class = {}
# sounds_noPack_HQ_per_class = {}
#
# # stage 1: display info about the pack status for every category to see where we are
# for cat_id, group_HQ in result_final_HQ.iteritems():
#
#     print('Analyzing packs in %s' % data_onto_by_id[cat_id]['name'])
#     # go through HQ sounds and extract pack info
#     count_noPack_HQ = 0
#     pack_status_HQ_per_class[cat_id] = {}
#     sounds_noPack_HQ_per_class[cat_id] = []
#     # pack_status_HQ[cat_id]['noPack_HQ'] = []
#
#     for fs_id in group_HQ:
#         # given a sound_id_target which pack we want to omit, retrieve pack_id
#         sound = client.get_sound(fs_id)
#
#         if sound.pack:
#             # sound belongs to a pack
#             pack_id = int(sound.pack.split('/')[-2])
#             if pack_id not in pack_status_HQ_per_class[cat_id]:
#                 # create pack if new pack for cat_id
#                 pack_status_HQ_per_class[cat_id][pack_id] = {}
#                 pack_status_HQ_per_class[cat_id][pack_id]['name'] = sound.pack_name
#                 pack_status_HQ_per_class[cat_id][pack_id]['fs_ids'] = []
#                 pack_status_HQ_per_class[cat_id][pack_id]['fs_ids'].append(fs_id)
#             else:
#                 # already existed. just append
#                 pack_status_HQ_per_class[cat_id][pack_id]['fs_ids'].append(fs_id)
#
#         else:
#             # sound DOES NOT belong to pack
#             sounds_noPack_HQ_per_class[cat_id].append(fs_id)
#             count_noPack_HQ += 1
#
#     # sanity check for the class: number of sounds before and after
#     if len(group_HQ) != (count_noPack_HQ + int(sum([len(value['fs_ids']) for key, value in pack_status_HQ_per_class[cat_id].iteritems()]))):
#         sys.error('PACK parsing error in %s' % data_onto_by_id[cat_id]['name'])
#
#
#
#
#
#







# client = freesound.FreesoundClient()
# client.set_token("eaa4f46407adf86c35c5d5796fd6ea8b05515dca", "token")

# given a sound_id_target which pack we want to omit, retrieve pack_id
# sound = client.get_sound(sound_id_target)
# pack_id = int(sound.pack.split('/')[-2])
# print pack_id, sound.pack_name

# retrieve the list of sound ids in the pack
# pack = client.get_pack(pack_id)
# pack_sounds = pack.get_sounds(fields="id,name,username", page_size=150)

# list_ids_pack_sounds = [sound.id for sound in pack_sounds]
# list_ids_pack_sounds.sort()

# confirm
# print 'all the %d sounds in the pack: %s, of user %s' % (len(list_ids_pack_sounds), pack.name, pack.username)
# for idx, sound in enumerate(pack_sounds):
#     print (idx + 1), sound.id, sound.name
#
# return list_ids_pack_sounds

















# stage 2: implement the split strategy that is more appropriate







# ---------------------------------------------------------------- #

# -------------------- REMOVE SOME CATEGORIES -------------------- #
print '\n FILTER CATEGORIES \n'

sounds_left = []
for idx in range(len(dataset_dev)):
    sounds_left += dataset_dev[idx]['sound_ids']
    sounds_left += dataset_eval[idx]['sound_ids']
print 'Number of sounds before filtering: {0}'.format(len(set(sounds_left)))

category_id_to_remove = ['/m/0c1dj', '/m/07phxs1', '/m/02rr_', '/m/07s0s5r', '/m/0l14qv', '/m/05jcn', '/m/025l19', '/m/01b9nn', '/m/01jnbd', '/m/05mxj0q', '/m/06mb1', '/m/02hnl', '/m/02zsn', '/m/07r660_', '/t/dd00093', '/m/01vfsf']

dataset_dev_filter = [d for d in dataset_dev if d['audioset_id'] not in category_id_to_remove]
dataset_eval_filter = [d for d in dataset_eval if d['audioset_id'] not in category_id_to_remove]

nb_labels_left = sum([len(d['sound_ids']) for d in dataset_dev_filter] + [len(d['sound_ids']) for d in dataset_eval_filter])

print 'Number of categories left: {0}'.format(len(dataset_dev_filter))

print 'Number of labels left: {0}'.format(nb_labels_left)

sounds_left = []
for idx in range(len(dataset_dev_filter)):
    sounds_left += dataset_dev_filter[idx]['sound_ids']
    sounds_left += dataset_eval_filter[idx]['sound_ids']

for aso_id in category_id_to_remove:
    try:
        del data_dev[aso_id]
        del data_dev_HQ[aso_id]
        del data_dev_LQ[aso_id]
        del data_dev_LQpior[aso_id]
        del data_eval[aso_id]
    except:
        pass
    
print 'Number of sounds left: {0}'.format(len(set(sounds_left)))


# ---------------------------------------------------------------- #

# -------------- REMOVE SOUND WITH MULTIPLE LABELS --------------- #
print '\n REMOVE MULTILABELED SOUNDS \n'
sound_to_remove = [s for s in sounds_left if sounds_left.count(s)>1]
for aso_id in data_dev.keys():
    for s in sound_to_remove:
        if s in data_dev[aso_id]:
            data_dev[aso_id].remove(s)
        if s in data_dev_HQ[aso_id]:
            data_dev_HQ[aso_id].remove(s)
        if s in data_dev_LQ[aso_id]:
            data_dev_LQ[aso_id].remove(s)
        if s in data_dev_LQpior[aso_id]:
            data_dev_LQpior[aso_id].remove(s)
        if s in data_eval[aso_id]:
            data_eval[aso_id].remove(s)

dataset_dev = [{'name': ontology_by_id[node_id]['name'], 
                'audioset_id': node_id,
                'sound_ids': data_dev[node_id],
               } for node_id in data_dev]

dataset_eval = [{'name': ontology_by_id[node_id]['name'], 
                'audioset_id': node_id,
                'sound_ids': data_eval[node_id],
               } for node_id in data_eval]
      

# ---------------------------------------------------------------- #

# --------------------- EXPORT DATASET JSON ---------------------- #
json.dump(data_dev_HQ, open(FOLDER_DATA + '/json/data_dev_HQ.json', 'w'))
json.dump(data_eval, open(FOLDER_DATA + '/json/data_eval.json', 'w'))


# ---------------------------------------------------------------- #

# --------------------- PRINT ALL CATEGORIES --------------------- #
### UTILS FUNCTIONS ###
def get_parents(aso_id, ontology):
    parents = []
    ontology_by_id = {o['id']:o for o in ontology}
    # 1st pass for direct parents
    for o in ontology:
        for id_child in o["child_ids"]:
            if id_child == aso_id:
                parents.append(o['id'])
    return [ontology_by_id[parent] for parent in parents]
        
def get_all_parents(aso_id, ontology): 
    """ 
    Recursive method to get the parents chain for an aso category
    """
    ontology_by_id = {o['id']:o for o in ontology}
    def paths(node_id, cur=list()):
        parents = get_parents(node_id, ontology)
        if not parents:
            yield cur
        else:
            for node in parents:
                for path in paths(node['id'], [node['name']] + cur):
                    yield path
    return paths(aso_id)

def sorted_occurrences_labels(data_dev_HQ, data_dev_LQ, data_dev_LQpior, data_dev, data_eval, ontology):
    """
    Create the worksheet with the number of sounds in each category of the ASO 
    Arguments:  - result from previous stage, e.g. result_leaves
                - ontology from json file
    """
    ontology_by_id = {o['id']:o for o in ontology}
    category_occurrences = []
    total_sounds = 0
    for node_id in data_dev_HQ.keys():
        nb_sample_dev_HQ = len(data_dev_HQ[node_id])
        nb_sample_dev_LQ = len(data_dev_LQ[node_id])
        nb_sample_eval = len(data_eval[node_id])
        nb_sample_dev = len(data_dev[node_id])
        nb_sample_dev_LQ_prior = len(data_dev_LQpior[node_id])
        total_sounds += nb_sample_dev + nb_sample_eval
        # get the names of parents (if several path, take one only and add (MULTIPLE PARENTS))
        all_parents = list(get_all_parents(node_id, ontology))
        if len(all_parents) > 1:
            names = ' > '.join(all_parents[0]+[ontology_by_id[node_id]['name']]) + ' (MULTIPLE PARENTS)'
        else:
            names = ' > '.join(all_parents[0]+[ontology_by_id[node_id]['name']])
        category_occurrences.append((names, node_id, nb_sample_dev_HQ, nb_sample_dev_LQ, nb_sample_dev_LQ_prior, nb_sample_dev, nb_sample_eval))
    category_occurrences = sorted(category_occurrences, key=lambda oc: oc[0])
    category_occurrences.append(('name', 'id', 'num dev HQ', 'num dev LQ', 'num dev LQ prior', 'num dev final', 'num eval'))
    category_occurrences.append(('Total number of sounds in final dataset', '', total_sounds, '', '', '', ''))
    category_occurrences.reverse()
    
    workbook = xlsxwriter.Workbook(FOLDER_DATA + 'list_categories_dataset_draft.xlsx')
    worksheet = workbook.add_worksheet('list categories')
    
    for idx, obj in enumerate(category_occurrences):
        worksheet.write(idx, 0, obj[0])
        worksheet.write(idx, 1, obj[1])
        worksheet.write(idx, 2, obj[2])
        worksheet.write(idx, 3, obj[3])
        worksheet.write(idx, 4, obj[4])
        worksheet.write(idx, 5, obj[5])
        worksheet.write(idx, 6, obj[6])
#    print '\n'
#    print 'Audio Set categories with their number of audio samples:\n'
#    for i in category_occurrences:
#        print str(i[0]).ljust(105) + str(i[1])


### WRITE EXCEL FILE ###
sorted_occurrences_labels(data_dev_HQ, data_dev_LQ, data_dev_LQpior, data_dev, data_eval, data_onto)


# --------------------------------------------------------------- #

# ---------------------- SPLIT LICENSE FILES -------------------- #
sound_ids_dev = set()
sound_ids_eval = set()
for category in dataset_dev:
    sound_ids_dev.update(category['sound_ids'])
for category in dataset_eval:
    sound_ids_eval.update(category['sound_ids'])
sound_ids_dev = list(sound_ids_dev)  
sound_ids_eval = list(sound_ids_eval)
sound_ids_dev.sort()
sound_ids_eval.sort()

license_file = open(FOLDER_DATA + 'licenses_dev.txt', 'w')
license_file.write("This dataset uses the following sounds from Freesound:\n\n")
license_file.write("to access user page:  http://www.freesound.org/people/<username>\n")
license_file.write("to access sound page: http://www.freesound.org/people/<username>/sounds/<soundid>\n\n")
license_file.write("'<file name>' with ID <soundid> by <username> [<license>]\n\n")
for sound_id in sound_ids_dev:
    sound = data_mapping[str(sound_id)]
    name = sound['name'].encode('utf-8').replace('\r', '')
    username = sound['username'].encode('utf-8').replace('\r', '')
    license_file.write("'{0}' with ID {1} by {2} [CC-{3}]\n"
                       .format(name, sound_id, username, sound['license'].split('/')[-3].upper()))
license_file.close()

license_file = open(FOLDER_DATA + 'licenses_eval.txt', 'w')
license_file.write("This dataset uses the following sounds from Freesound:\n\n")
license_file.write("to access user page:  http://www.freesound.org/people/<username>\n")
license_file.write("to access sound page: http://www.freesound.org/people/<username>/sounds/<soundid>\n\n")
license_file.write("'<file name>' with ID <soundid> by <username> [<license>]\n\n")
for sound_id in sound_ids_eval:
    sound = data_mapping[str(sound_id)]
    name = sound['name'].encode('utf-8').replace('\r', '')
    username = sound['username'].encode('utf-8').replace('\r', '')
    license_file.write("'{0}' with ID {1} by {2} [CC-{3}]\n"
                       .format(name, sound_id, username, sound['license'].split('/')[-3].upper()))
license_file.close()


# --------------------------------------------------------------- #

# -------------------------- CREATE CSV ------------------------- #
try:
    merge = json.load(open(FOLDER_DATA + 'json/merge_categories.json', 'rb'))
except:
    raise Exception('CREATE THE FILE "merge_categories.json"')
node_id_parent = {}
for d in merge:
    for dd in merge[d]:
        node_id_parent[dd] = d
        
#ontology = json.load(open('ontology/ontology.json', 'rb'))
#ontology_by_id = {o['id']:o for o in ontology}
sounds_A = [] # sounds for dataset A
sounds_B = [] # sounds for dataset B

import csv
with open(FOLDER_DATA + 'dataset_dev.csv', 'wb') as f:
    writer = csv.writer(f)
    for d in dataset_dev:
        for sound_id in d['sound_ids']:
            if sound_id in data_dev_HQ[d['audioset_id']]:
                quality = 1
            elif sound_id in data_dev_LQ[d['audioset_id']] or sound_id in data_dev_LQpior[d['audioset_id']]:
                quality = 0
            sounds_A.append((sound_id, data_mapping[str(sound_id)]['duration']))
            try:
                writer.writerow([sound_id, d['audioset_id'], d['name'], node_id_parent[d['audioset_id']], ontology_by_id[node_id_parent[d['audioset_id']]]['name'], quality])
                sounds_B.append((sound_id, data_mapping[str(sound_id)]['duration']))
            except:
                writer.writerow([sound_id, d['audioset_id'], d['name'], None, None, quality])

with open(FOLDER_DATA + 'dataset_eval.csv', 'wb') as f:
    writer = csv.writer(f)
    for d in dataset_eval:
        for sound_id in d['sound_ids']:
            sounds_A.append((sound_id, data_mapping[str(sound_id)]['duration']))
            try:
                writer.writerow([sound_id, d['audioset_id'], d['name'], node_id_parent[d['audioset_id']], ontology_by_id[node_id_parent[d['audioset_id']]]['name']])
                sounds_B.append((sound_id, data_mapping[str(sound_id)]['duration']))
            except:
                writer.writerow([sound_id, d['audioset_id'], d['name'], None, None])

print 'Total duration of the dataset A: {0} secondes'.format(sum([s[1] for s in list(set(sounds_A))]))
print 'Total duration of the dataset B: {0} secondes'.format(sum([s[1] for s in list(set(sounds_B))]))
print 'Total of classes in dataset B: {0}'.format(len(merge))

all_ids = [s[0] for s in sounds_A]
json.dump(all_ids, open(FOLDER_DATA + 'all_freesound_ids.json', 'w'))
            
# --------------------------------------------------------------- #
